#!/usr/bin/env python3
"""
MUI Translation Tool — GUI Version
雙擊執行，選 zip + Excel，按 Run 即可。
"""

import json, re, sys, zipfile, threading, heapq, os
from pathlib import Path
from collections import defaultdict, OrderedDict, Counter
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

APP_VERSION  = "v2.2.BUILD_DATETIME"   # replaced by build script at package time
APP_AUTHOR   = "Pocky Wu"
TOOL_VERSION = "6"   # bump when index structure changes (forces cache rebuild)

CHANGELOG = """\
v2.2
────────────────────────────────────────
新功能
  • 新增「YCA」App（YouCam AI Pro，16 語言）
  • YCVB 新增 HEB 希伯來語、ARA 阿拉伯語（共 18 語言）

────────────────────────────────────────
v2.1
────────────────────────────────────────
Bug 修正
  • 切換模式時「轉換 Ignore」按鈕在 Web App 下會短暫恢復正常樣式

新功能
  • YCVB 新增 ARA 阿拉伯語、HEB 希伯來語（共 18 語言）
  • 新增「Web」App 支援（cosmetic-web-doc JSON 格式）：
    → 動態偵測 Zip 內語言，無需手動設定
    → 輸出 Excel 無 Module 欄（Web 翻譯無模組概念）
    → 支援 Excel 查詢、語言全掃描、比對新字串三個模式
    → 不支援「轉換 Ignore」模式（Web 無對應掃描報告格式）
  • 新增 5 種語言：VIE 越南語 / HIN 印地語 / TAM 泰米爾語 / FAR 波斯語 / AFR 南非荷蘭語

────────────────────────────────────────
v2.0
────────────────────────────────────────
新功能
  • 新增「比對新字串」模式：
    → 選取舊版與新版 Zip，一鍵輸出新版新增的英文字串
    → 輸出 Excel（Sheet: 新增字串）含 Module / New Strings (EN) / Key 三欄，按 Module 排序分組
    → 自動加 Auto Filter、凍結首列、交替列色
    → 四個模式路徑一致：選檔後自動記憶、切換 App 自動 reset/restore、
      輸出檔已存在跳出四選一彈窗（與其他模式行為完全同步）

新功能
  • Test Sheet 新增「自定」欄：manager 可對特定語言從下拉選單指定 Tester
    → 下拉內容自動對應 M3:M8 姓名，有填則覆蓋自動分配，空白則繼續用 Greedy

修正
  • YCE 匈牙利語代碼 HUM → HUN（與 YCP 及 ISO 639-2 標準統一）

UI
  • 模式按鈕改為 2×2 等寬等高 Grid 排版，各按鈕間保留 3px 間距
  • 比對新字串模式右側改為功能說明卡，不再顯示無關的快速查詢面板
  • 四個執行按鈕統一改為藍色（#89b4fa）

────────────────────────────────────────
v1.9
────────────────────────────────────────
原有功能
  • 報告新增「Test Sheet」工作表：
    → 只列有黃色標記（最長字串）的語言與字串，相同翻譯去重
    → 查無字串（Server string）自動展開成每語言一列（amber 底色）
    → 語言群組交替藍 / 白底色，視覺上易區分語言切換
    → Tester 自動分配：Greedy bin-packing 預算 n=2~6 人最佳分配
      改 M2 人數 → Tester 欄自動切換；改 M 欄姓名 → 即時更新
    → 欄位：Language / Page / Module / EN (Base) / Translation / Tester / Test result
    → 右側統計表顯示每語言字串數 + Total，支援 Auto Filter
  • YCP 新增 7 種語言：UKR / BUL / SWE / HRV / HUN / DAN / GRN
    → YCP 語言總數 20 → 27

Bug 修正
  • YCP 波蘭語代碼 PLO 修正為 POL

────────────────────────────────────────
v1.8
────────────────────────────────────────
新功能
  • 快速查詢 & Excel 模式皆支援 * 萬用字元
    → 輸入 Premium plan* 自動展開所有開頭符合的字串，各自輸出翻譯結果
    → Excel 模式：一列萬用字元展開為多列，方便批次比對

強化 / Bug 修正
  • 模糊比對新增兩種正規化規則，命中率大幅提升：
    → [[in %1$d days]] 型 placeholder 剝除 [[ ]] 外殼後再比對
      （app 顯示的實際值如「in 7 days」可正確命中翻譯檔）
    → 裝飾用單引號 'History' → History，與 app 顯示一致

UI
  • 全介面配色改為 Catppuccin Mocha 主題（深色、對比層次分明）
  • 選取中的 App / 模式按鈕改用深色文字（#1e1e2e），確保淺紫背景上可讀
  • Log / 查詢結果框選取色改為 #313244，消除系統亮藍高亮條

────────────────────────────────────────
v1.7
────────────────────────────────────────
新功能
  • 新增「轉換 Ignore」模式：一鍵將語言掃描報告（含 <語言> Extracted Sheet）轉換為 Ignore Excel
    → Comment = Pass 的字串按 Sheet 所屬語言收錄，不混用「語言欄」的多語言值
    → 同一 Key 在多個語言都 Pass 時自動合併為同一列（語言欄逗號分隔）
    → 多個 Key 同格（換行分隔）自動拆分為獨立列
  • 所有檔案欄位加上 Hover Tooltip，滑入顯示完整路徑，三個模式皆支援

強化 / Bug 修正
  • 輸出檔已存在時跳出四選一彈窗：取代 / 自動加序號 / 取新名稱 / 取消
    → 三個模式（Excel 查詢、語言全掃描、轉換 Ignore）皆套用
  • 轉換 Ignore 語言 Sheet 偵測改用 Sheet 名稱大寫字母比對
    → FIL Extracted、FIL_Extraction 等命名變體皆可正常執行

UI
  • 全介面字型改為 Microsoft JhengHei UI（微軟正黑體 UI），中英文顯示更一致

────────────────────────────────────────
v1.6
────────────────────────────────────────
UI 優化
  • 語言全掃描模式：輸入／輸出 Excel 欄位自動灰化，避免使用者混淆
  • 語言全掃描模式：輸出報告「選取」按鈕不再被截斷（改用彈性欄寬）
  • 移除「部分 Key 未翻譯」分類，改統一顯示為「未翻譯 / 空白」，避免使用者混淆

新功能
  • Ignore Excel 語言欄支援 ALL_LANGUAGES，表示該字串所有語言皆與英文相同，直接跳過不報錯

Bug 修正
  • 修正特定翻譯字串含 \x10 等控制字元導致輸出 Excel 時崩潰（IllegalCharacterError）
    → 寫入前自動清除 Excel XML 不合法字元，不影響翻譯內容完整度

────────────────────────────────────────
v1.5
────────────────────────────────────────
新功能
  • 新增 YCP、YCVB 兩個 App（YCP 支援 20 語言含希臘/波蘭/羅馬尼亞，YCVB 16 語言）
  • Android XML parser 新增 <plurals> 支援，以 quantity="other" 作為代表值

UI 優化
  • App 選擇按鈕格式改為 YMK(20)，縮短寬度
  • 輸出 Excel 檔名欄寬加大（32 字元），減少截斷

────────────────────────────────────────
v1.4
────────────────────────────────────────
新功能
  • 更新記錄：Header 新增「📋 更新記錄」入口，可查看每版變更
  • 自動發佈腳本：build 完可一鍵上傳 exe 至 Google Drive 並更新版號
  • 自動更新提示：app 啟動時背景檢查是否有新版（版號格式改為 1.4.yyMMdd_HHmm）

UI 優化
  • 全介面字型放大兩號
  • Log 區域撐滿視窗底部，不再留白也不出現多餘捲動條
  • 檔名過長時截斷顯示，前綴 … 保留尾端資訊
  • 欄位標籤加冒號，「Zip 檔」改為「翻譯 Zip 檔:」
  • 語言勾選區塊字型縮小兩號、改為 4 欄，語言代碼完整顯示不截斷

Bug 修正
  • 修正 fuzzy match 時翻譯與英文相同未標紅（如 Choose 1-{{count}} styles）
  • 修正「輸出報告」選取鈕超出視窗右邊界

建置
  • build_windows.bat 改用 PowerShell 取時間戳，修正 Windows 11 移除 wmic 後無法 build 的問題

────────────────────────────────────────
v1.3
────────────────────────────────────────
新功能
  • 取消按鈕：Run / 開始掃描執行中可隨時中止
  • 進度條百分比：顯示「處理第 N / M」並同步更新進度條
  • 快速查詢獨立結果框：查詢結果不再混入 log，右欄獨立顯示
  • 語言全掃描記憶：切換 App 或重開程式自動帶入上次勾選的語言
  • 掃描輸出路徑記憶：與 Zip / Excel 路徑同步儲存於 config
  • 自動更新提示：啟動時背景檢查 Google Drive 是否有新版

Bug 修正
  • 修正 Android gen/ 資料夾的中文字串被誤當英文 base
  • 修正 .strings 檔解析 regex typo 導致部分字串漏抓
  • 修正 fuzzy match 情況下翻譯與英文相同未標紅（Choose 1-{{count}} styles 等）
  • 修正語言全掃描面板「輸出報告」選取鈕超出視窗邊界
  • 修正捲動條在內容未超出視窗時仍可捲動的問題

效能 / 架構
  • 索引快取邏輯重構，三個模式共用同一 _load_index()
  • _normalize() 改用 module-level 預編譯 regex，加速重複呼叫
  • UI 操作全面改用 self.after() 確保 thread safety

────────────────────────────────────────
v1.2
────────────────────────────────────────
  • 新增語言全掃描模式（不需 Excel，直接掃描整包 Zip）
  • 新增快速查詢（右側文字框，Ctrl+Enter 執行）
  • 新增 Ignore Excel 支援（跳過預期與英文相同的字串）
  • 翻譯問題報告新增「部分 Key 未翻譯」橘色分類
  • 模糊比對支援（數字 ↔ 變數 正規化，橘色字標示）
  • 路徑設定自動記憶，切換 App 各自保留上次路徑
"""

# ── Auto-update ───────────────────────────────────────────────────────────────
# latest_version.txt on Google Drive (public, "Anyone with link can view")
# Content format: full version with build stamp, e.g.  1.3.260508_1030
_UPDATE_CHECK_URL    = "https://drive.google.com/uc?export=download&id=11dRxedUqKwtD26P4EcNK17qmICefJ4-w"
_UPDATE_DOWNLOAD_URL = "https://drive.google.com/drive/folders/1-DumwCUoqOV-zgMRhUtI4QRMtqGn3qDq"
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    import pandas as pd
except ImportError:
    import subprocess, sys
    root = tk.Tk(); root.withdraw()
    if messagebox.askyesno("缺少套件", "需要安裝 openpyxl 和 pandas，是否立即安裝？"):
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "pandas"])
        messagebox.showinfo("完成", "安裝完成，請重新啟動程式。")
    sys.exit(0)

# ┌─────────────────────────────────────────────────────────────────────────┐
# │  APP_CONFIGS — 每個 App 的語言設定                                        │
# │  新增 App:  在下面複製一個 block，改名稱和語言清單                          │
# │  新增語言:  在對應 App 的 dict 裡加一行 "代碼": "lang code"                │
# │  設為 None = 輸出全部語言（不過濾）                                        │
# └─────────────────────────────────────────────────────────────────────────┘
APP_CONFIGS: dict[str, dict | None] = {

    "YMK": {
        "ENU": "en",
        "CHT": "zh-Hant",
        "CHS": "zh-Hans",
        "JPN": "ja",
        "KOR": "ko",
        "ESP": "es",
        "ARA": "ar",
        "FRA": "fr",
        "DEU": "de",
        "RUS": "ru",
        "PTB": "pt",
        "IND": "id",
        "THA": "th",
        "TRK": "tr",
        "MSL": "ms",
        "ITA": "it",
        "NLD": "nl",
        "HEB": "he",
        "POL": "pl",
        "FIL": "fil",
    },

    "YCE": {
        "ENU": "en",
        "CHT": "zh-Hant",
        "CHS": "zh-Hans",
        "JPN": "ja",
        "KOR": "ko",
        "RUS": "ru",
        "DEU": "de",
        "FRA": "fr",
        "THA": "th",
        "ESP": "es",
        "PTB": "pt",
        "IND": "id",
        "TRK": "tr",
        "MSL": "ms",
        "NLD": "nl",
        "ITA": "it",
        "HUN": "hu",
        "SWE": "sv",
        "ARA": "ar",
    },

    "YCP": {
        "ENU": "en",
        "CHT": "zh-Hant",
        "CHS": "zh-Hans",
        "JPN": "ja",
        "KOR": "ko",
        "DEU": "de",
        "ESP": "es",
        "FRA": "fr",
        "ITA": "it",
        "RUS": "ru",
        "PTB": "pt",
        "IND": "id",
        "THA": "th",
        "TRK": "tr",
        "MSL": "ms",
        "NLD": "nl",
        "ELL": "el",
        "POL": "pl",
        "RON": "ro",
        "ARA": "ar",
        "UKR": "uk",
        "BUL": "bg",
        "SWE": "sv",
        "HRV": "hr",
        "HUN": "hu",
        "DAN": "da",
        "GRN": "gn",
    },

    "YCVB": {
        "ENU": "en",
        "CHT": "zh-Hant",
        "CHS": "zh-Hans",
        "JPN": "ja",
        "KOR": "ko",
        "DEU": "de",
        "ESP": "es",
        "FRA": "fr",
        "ITA": "it",
        "RUS": "ru",
        "PTB": "pt",
        "IND": "id",
        "THA": "th",
        "TRK": "tr",
        "MSL": "ms",
        "NLD": "nl",
        "HEB": "he",
        "ARA": "ar",
    },

    "YCA": {
        "ENU": "en",
        "CHT": "zh-Hant",
        "CHS": "zh-Hans",
        "JPN": "ja",
        "KOR": "ko",
        "DEU": "de",
        "ESP": "es",
        "FRA": "fr",
        "ITA": "it",
        "RUS": "ru",
        "PTB": "pt",
        "IND": "id",
        "THA": "th",
        "TRK": "tr",
        "MSL": "ms",
        "NLD": "nl",
    },

    # Web: dynamic — languages detected from zip at runtime
    "Web": {},

}

REGIONAL_ALLOWLIST: set = set()

# ── Web locale mapping ────────────────────────────────────────────────────────
# locale filename stem → internal code  (e.g. "ar_AE" → "ARA")
_WEB_LOCALE_TO_CODE: dict[str, str] = {
    "en_US": "ENU",  "zh_TW": "CHT",  "zh_CN": "CHS",  "ja_JP": "JPN",
    "ko_KR": "KOR",  "ar_AE": "ARA",  "fr_FR": "FRA",  "de_DE": "DEU",
    "ru_RU": "RUS",  "pt_BR": "PTB",  "id_ID": "IND",  "th_TH": "THA",
    "tr_TR": "TRK",  "ms_MY": "MSL",  "it_IT": "ITA",  "nl_NL": "NLD",
    "he_IL": "HEB",  "pl_PL": "POL",  "fil_PH": "FIL", "hu_HU": "HUN",
    "sv_SE": "SWE",  "el_GR": "ELL",  "ro_RO": "RON",  "uk_UA": "UKR",
    "bg_BG": "BUL",  "hr_HR": "HRV",  "da_DK": "DAN",  "gn_PY": "GRN",
    "vi_VN": "VIE",  "hi_IN": "HIN",  "ta_IN": "TAM",  "fa_IR": "FAR",
    "af_ZA": "AFR",
}
# locale filename stem → ISO lang code  (e.g. "ar_AE" → "ar")
_WEB_LOCALE_TO_LANG: dict[str, str] = {
    "en_US": "en",      "zh_TW": "zh-Hant", "zh_CN": "zh-Hans", "ja_JP": "ja",
    "ko_KR": "ko",      "ar_AE": "ar",       "fr_FR": "fr",       "de_DE": "de",
    "ru_RU": "ru",      "pt_BR": "pt",        "id_ID": "id",       "th_TH": "th",
    "tr_TR": "tr",      "ms_MY": "ms",        "it_IT": "it",       "nl_NL": "nl",
    "he_IL": "he",      "pl_PL": "pl",        "fil_PH": "fil",     "hu_HU": "hu",
    "sv_SE": "sv",      "el_GR": "el",        "ro_RO": "ro",       "uk_UA": "uk",
    "bg_BG": "bg",      "hr_HR": "hr",        "da_DK": "da",       "gn_PY": "gn",
    "vi_VN": "vi",      "hi_IN": "hi",        "ta_IN": "ta",       "fa_IR": "fa",
    "af_ZA": "af",
}

def _is_web_zip(zip_path: Path) -> bool:
    """Return True if zip contains top-level en_US.json (web translation format)."""
    try:
        with zipfile.ZipFile(zip_path) as zf:
            return "en_US.json" in zf.namelist()
    except Exception:
        return False

def _detect_web_langs_from_zip(zip_path: Path) -> dict:
    """
    Fast scan of zip filenames → {internal_code: lang_code} for all detected locales.
    Does NOT read file contents — used for UI updates only.
    """
    result = {}
    try:
        with zipfile.ZipFile(zip_path) as zf:
            names = {Path(n).name for n in zf.namelist()}
        for stem, code in _WEB_LOCALE_TO_CODE.items():
            if stem == "en_US": continue
            if stem + ".json" in names:
                result[code] = _WEB_LOCALE_TO_LANG[stem]
    except Exception:
        pass
    return result

# ── Language metadata ─────────────────────────────────────────────────────────
LANG_NAMES = {
    "en":      "英語",
    "zh-Hant": "繁體中文",  "zh-Hans": "簡體中文",  "zh": "中文",
    "ja": "日語",           "ko": "韓語",           "ar": "阿拉伯語",
    "de": "德語",           "es": "西班牙語",        "fr": "法語",
    "it": "義大利語",       "nl": "荷蘭語",          "pt": "葡萄牙語",
    "pt-PT": "葡萄牙語(葡)", "ru": "俄語",           "th": "泰語",
    "tr": "土耳其語",       "id": "印尼語",          "ms": "馬來語",
    "fil": "菲律賓語",      "he": "希伯來語",        "pl": "波蘭語",
    "sv": "瑞典語",         "da": "丹麥語",          "hu": "匈牙利語",
    "el": "希臘語",         "uk": "烏克蘭語",        "bg": "保加利亞語",
    "hr": "克羅埃西亞語",   "ro": "羅馬尼亞語",      "cs": "捷克語",
    "gn": "瓜拉尼語",
    "vi": "越南語",   "hi": "印地語",
    "ta": "泰米爾語", "fa": "波斯語",
    "af": "南非荷蘭語",
}
LANG_PRIORITY = ["en", "zh-Hant", "zh-Hans", "zh", "ja", "ko"]

# ── Styles ────────────────────────────────────────────────────────────────────
YELLOW  = PatternFill("solid", fgColor="FFD700")
YELLOW2 = PatternFill("solid", fgColor="FFF176")   # 2nd longest when ARA is longest
RED     = PatternFill("solid", fgColor="FF6B6B")
HEADER = PatternFill("solid", fgColor="302B63")
H_FONT = Font(bold=True, color="FFFFFF", name="Microsoft JhengHei UI", size=10)
D_FONT = Font(name="Microsoft JhengHei UI", size=10)
B_FONT = Font(name="Microsoft JhengHei UI", size=10, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
thin   = Side(style="thin", color="DDDDDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Config persistence (#6) ───────────────────────────────────────────────────
import tempfile as _tempfile

# Store config in a stable OS-appropriate location (not tempdir which can be wiped)
if sys.platform == "win32":
    _cfg_base = Path(os.environ.get("APPDATA", str(Path.home())))
elif sys.platform == "darwin":
    _cfg_base = Path.home() / "Library" / "Application Support"
else:
    _cfg_base = Path(os.environ.get("XDG_CONFIG_HOME", str(Path.home() / ".config")))
_CONFIG_PATH     = _cfg_base / "TranslationTool" / "config.json"
_OLD_CONFIG_PATH = Path(_tempfile.gettempdir()) / "TranslationTool" / "config.json"

def load_config() -> dict:
    try:
        if _CONFIG_PATH.exists():
            return json.loads(_CONFIG_PATH.read_text("utf-8"))
        # Migrate from old tempdir location (one-time)
        if _OLD_CONFIG_PATH.exists():
            data = json.loads(_OLD_CONFIG_PATH.read_text("utf-8"))
            save_config(data)   # write to new location immediately
            return data
    except Exception:
        pass
    return {}

def save_config(cfg: dict):
    try:
        _CONFIG_PATH.parent.mkdir(exist_ok=True)
        _CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass

# ── Parsers ───────────────────────────────────────────────────────────────────

def _is_regional_variant(lang: str) -> bool:
    if lang in REGIONAL_ALLOWLIST: return False
    m = re.match(r'^[a-z]{2,3}-([A-Z][a-z]{3}|[A-Z]{2})$', lang)
    if not m: return False
    return len(m.group(1)) == 2

def parse_arb(text: str) -> tuple:
    try: data = json.loads(text)
    except: return "", {}
    loc  = data.get("@@locale") or data.get("localeCode")
    lang = loc.replace("_", "-") if loc else ""
    return lang, {k: v for k, v in data.items()
                  if not k.startswith("@") and isinstance(v, str)}

def parse_strings(text: str) -> dict:
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)
    entries = {}
    for line in text.splitlines():
        s = line.strip()
        if not s or s.startswith('//'): continue
        m = re.match(r'^"((?:[^"\\]|\\.)*)"[ \t]*=[ \t]*"((?:[^"\\]|\\.)*)"\s*;', s)
        if m: entries[m.group(1)] = m.group(2)
    return entries

def lang_from_arb_name(stem: str) -> str:
    m = re.search(r'_([a-z]{2}(?:_[A-Za-z]{2,4})*)$', stem)
    return m.group(1).replace("_", "-") if m else ""

def lang_from_strings_name(name: str) -> str:
    m = re.match(r'.+?_(.+?)(?:\.lproj)?\.strings$', name)
    return m.group(1).replace("_", "-") if m else ""

ANDROID_LANG_MAP = {
    "zh-rCN": "zh-Hans", "zh-rHK": "zh-Hant", "zh-rTW": "zh-Hant",
    "in": "id", "iw": "he", "pt-rBR": "pt", "nl-rNL": "nl", "nl-rBE": "nl",
}

def lang_from_android_folder(folder: str) -> str:
    if folder == "values": return "en"
    m = re.match(r'^values-(.+)$', folder)
    if not m: return ""
    raw = m.group(1)
    if re.match(r'^v\d+$', raw): return ""
    if raw in ANDROID_LANG_MAP: return ANDROID_LANG_MAP[raw]
    return re.sub(r'-r([A-Z]{2})$', r'-\1', raw)

def parse_android_xml(text: str) -> dict:
    import html as _html

    def _clean(raw: str) -> str:
        raw = raw.strip()
        raw = raw.replace("\\'", "'").replace('\\"', '"')
        raw = raw.replace("\\n", "\n").replace("\\t", "\t")
        raw = re.sub(r'<!\[CDATA\[(.*?)\]\]>', r'\1', raw, flags=re.DOTALL)
        raw = re.sub(r'<[^>]+>', '', raw)
        return _html.unescape(raw).strip()

    entries = {}
    for m in re.finditer(r'<string\s+name="([^"]+)"[^>]*>(.*?)</string>', text, re.DOTALL):
        val = _clean(m.group(2))
        if m.group(1) and val:
            entries[m.group(1)] = val

    # <plurals>: use "other" quantity as representative value
    for m in re.finditer(r'<plurals\s+name="([^"]+)"[^>]*>(.*?)</plurals>', text, re.DOTALL):
        key, block = m.group(1), m.group(2)
        item = re.search(r'<item\s+quantity="other"[^>]*>(.*?)</item>', block, re.DOTALL)
        if not item:
            item = re.search(r'<item[^>]*>(.*?)</item>', block, re.DOTALL)
        if item:
            val = _clean(item.group(1))
            if key and val:
                entries[key] = val

    return entries

GARBLED_RE = re.compile(r'[\ufffd\x00-\x08\x0b\x0c\x0e-\x1f]')

# Characters illegal in Excel XML (openpyxl will raise IllegalCharacterError)
_XL_ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\ud800-\udfff]')

def _xl_safe(s: str) -> str:
    """Strip characters that openpyxl cannot write to Excel cells."""
    if not isinstance(s, str): return s
    return _XL_ILLEGAL_RE.sub('', s)

def is_garbled(val: str) -> bool:
    if not val or not val.strip(): return True
    if GARBLED_RE.search(val): return True
    if len(val) > 3 and val.count('?') > len(val) * 0.4: return True
    return False

def _visual_len(s: str) -> int:
    """
    Count visible characters only, excluding Unicode combining marks
    (e.g. Thai tone marks ่ ้ ๊ ๋, Arabic diacritics ً ٌ ٍ etc.)
    so that 'หลีกเลี่ยง' and 'Vermeiden' compare by display width.
    """
    import unicodedata
    return sum(1 for c in s if unicodedata.category(c) not in ('Mn', 'Mc', 'Me'))


_RE_PRINTF   = re.compile(r'%\d+\$[@disfeEgGuoxXld]+')
_RE_PRINTF2  = re.compile(r'%[@disfeEgGuoxXld]+')
_RE_BRACE    = re.compile(r'\{[^}]+\}')
_RE_DBRACKET = re.compile(r'\[\[|\]\]')   # strip [[ ]] wrappers, keep inner content
_RE_NUM      = re.compile(r'(?<!\w)\d+(?!\w)')
_RE_SPACE    = re.compile(r'\s+')

def _normalize(s: str) -> str:
    s = s.replace('\\n', ' ').replace('\n', ' ')
    s = _RE_PRINTF.sub('__X__', s)
    s = _RE_PRINTF2.sub('__X__', s)
    s = _RE_BRACE.sub('__X__', s)
    s = _RE_DBRACKET.sub('', s)           # [[in %1$d days]] → in %1$d days
    s = s.replace("'", "")               # 'History' → History
    s = _RE_NUM.sub('__X__', s)
    return _RE_SPACE.sub(' ', s).strip().lower()

# ── Index builder ─────────────────────────────────────────────────────────────

def build_index(source: Path, base_lang: str, log) -> tuple:
    log(f"🔍 掃描翻譯來源: {source.name}")
    projects: dict = defaultdict(dict)

    def add(proj, lang, kvs):
        if proj and lang and kvs:
            if lang in projects[proj]: projects[proj][lang].update(kvs)
            else: projects[proj][lang] = dict(kvs)

    def process_file(proj, fname, raw, folder=""):
        if fname.endswith(".arb"):
            lc, kvs = parse_arb(raw)
            lang = lc or lang_from_arb_name(Path(fname).stem)
            add(proj, lang, kvs)
        elif fname.endswith(".strings"):
            add(proj, lang_from_strings_name(fname), parse_strings(raw))
        elif fname == "strings.xml" and folder:
            lang = lang_from_android_folder(folder)
            if lang: add(proj, lang, parse_android_xml(raw))

    if zipfile.is_zipfile(source):
        with zipfile.ZipFile(source) as zf:
            for info in zf.infolist():
                if info.is_dir(): continue
                if "/gen/" in info.filename.replace("\\", "/"): continue
                parts = Path(info.filename).parts
                if len(parts) < 2: continue
                fname, folder = parts[-1], parts[-2]
                raw = zf.read(info).decode("utf-8", errors="replace")
                if fname == "strings.xml":
                    proj = parts[1] if len(parts) >= 7 else parts[0]
                    process_file(proj, fname, raw, folder)
                else:
                    process_file(parts[-2], fname, raw)
    elif source.is_dir():
        for pd_ in sorted(source.iterdir()):
            if not pd_.is_dir(): continue
            for fp in sorted(pd_.rglob("*")):
                if fp.is_dir(): continue
                try: raw = fp.read_text("utf-8", errors="replace")
                except: continue
                process_file(pd_.name, fp.name, raw, fp.parent.name)

    lookup: dict = {}
    for proj, lang_data in projects.items():
        base_kvs = lang_data.get(base_lang, {})
        if not base_kvs: continue

        # First pass: group all keys by their en_val
        # key_groups[en_val] = [ (key, {lang: val}) ]
        key_groups: dict[str, list] = {}
        for key, en_val in base_kvs.items():
            if not en_val.strip(): continue
            pt = {}
            for lang, kvs in lang_data.items():
                if lang == base_lang: continue
                val = kvs.get(key, "")
                if val: pt[lang] = val
            key_groups.setdefault(en_val, []).append((key, pt))

        # Second pass: merge per en_val
        for en_val, key_list in key_groups.items():
            # All langs that exist in any key's translations
            all_langs = set()
            for _, pt in key_list:
                all_langs.update(pt.keys())

            # Best translation per lang (from whichever key has it)
            best: dict = {base_lang: en_val}
            for lang in all_langs:
                for _, pt in key_list:
                    val = pt.get(lang, "")
                    if val:
                        best[lang] = val
                        break

            # Track which keys are missing each lang
            key_missing: dict[str, list] = {}   # lang → [missing key names]
            if len(key_list) > 1:
                # Only worth tracking when multiple keys share the same en_val
                for lang in all_langs:
                    missing_keys = [k for k, pt in key_list if not pt.get(lang)]
                    if missing_keys:
                        key_missing[lang] = missing_keys

            # Also flag langs that are missing from ALL keys
            for lang in all_langs:
                if not best.get(lang):
                    key_missing.setdefault(lang, [k for k, _ in key_list])

            best["_key_missing"] = key_missing
            best["_all_keys"]    = [k for k, _ in key_list]
            best["_key_trans"]   = {k: dict(pt) for k, pt in key_list}  # per-key translations
            lookup.setdefault(en_val, {})[proj] = best

    norm_index: dict = {}
    for en_val in lookup:
        nk = _normalize(en_val)
        if nk not in norm_index:
            norm_index[nk] = en_val

    log(f"✅ 索引完成: {len(projects)} 個 project，{len(lookup)} 個英文字串")
    return lookup, norm_index

# ── Web index builder ─────────────────────────────────────────────────────────

def build_web_index(source: Path, log) -> tuple:
    """
    Build index from a web-format Zip (flat {key: value} JSON files per locale).
    Returns (lookup, norm_index, detected_langs) where:
      detected_langs = {internal_code: lang_code}  e.g. {"ARA": "ar", "FRA": "fr"}
    The lookup uses "" as pseudo-module (web has no module concept).
    """
    log(f"🔍 掃描 Web 翻譯來源: {source.name}")

    with zipfile.ZipFile(source) as zf:
        all_names = {Path(n).name: n for n in zf.namelist()}
        en_data: dict = json.loads(zf.read(all_names["en_US.json"]).decode("utf-8"))

        lang_data: dict = {}   # lang_code → {key: value}
        detected: dict  = {}   # internal_code → lang_code

        for stem, code in _WEB_LOCALE_TO_CODE.items():
            if stem == "en_US": continue
            fname = stem + ".json"
            if fname not in all_names: continue
            lc = _WEB_LOCALE_TO_LANG[stem]
            lang_data[lc] = json.loads(zf.read(all_names[fname]).decode("utf-8"))
            detected[code] = lc

    # Group keys by EN value (same EN value may map to multiple keys)
    key_groups: dict[str, list] = {}
    for key, en_val in en_data.items():
        if not en_val.strip(): continue
        translations: dict = {}
        for lc, kv in lang_data.items():
            val = kv.get(key, "")
            if val: translations[lc] = val
        key_groups.setdefault(en_val, []).append((key, translations))

    lookup: dict = {}
    for en_val, key_list in key_groups.items():
        all_langs: set = set()
        for _, pt in key_list: all_langs.update(pt.keys())

        best: dict = {"en": en_val}
        for lang in all_langs:
            for _, pt in key_list:
                val = pt.get(lang, "")
                if val: best[lang] = val; break

        best["_all_keys"]    = [k for k, _ in key_list]
        best["_key_trans"]   = {k: dict(pt) for k, pt in key_list}
        best["_key_missing"] = {}

        # "" = pseudo-module (no module in web format)
        lookup.setdefault(en_val, {})[""] = best

    norm_index = {_normalize(k): k for k in lookup}
    log(f"✅ Web 索引完成: {len(lookup)} 個英文字串，{len(detected)} 個語言")
    return lookup, norm_index, detected


# ── Ignore list loader ────────────────────────────────────────────────────────

def load_ignore_list(path: Path, target_langs: dict | None = None) -> set:
    """
    Parse ignore Excel with columns: Module | ENU翻譯檔字串 | 問題Key | 語言
    語言 column contains comma-separated lang codes (e.g. "HEB, POL, FIL")
    using the APP_CONFIGS keys (e.g. HEB→he, POL→pl, FIL→fil).
    Special value "ALL_LANGUAGES" skips the string for every language.
    Returns set of (module, key_name, lang_code) to skip in reports.
    Entries with ALL_LANGUAGES use lang_code="*" as a wildcard sentinel.
    """
    if not path or not path.exists():
        return set()

    import pandas as pd_
    try:
        df = pd_.read_excel(str(path), header=0)
    except Exception:
        return set()

    # Build code→lang mapping e.g. {"HEB": "he", "POL": "pl"}
    code_to_lang = {k: v for k, v in (target_langs or {}).items()}

    ignore: set = set()
    for _, row in df.iterrows():
        module    = str(row.iloc[0]).strip() if pd_.notna(row.iloc[0]) else ""
        key_name  = str(row.iloc[2]).strip() if pd_.notna(row.iloc[2]) else ""
        langs_str = str(row.iloc[3]).strip() if pd_.notna(row.iloc[3]) else ""

        if not module or not key_name or not langs_str:
            continue

        if langs_str.upper() == "ALL_LANGUAGES":
            ignore.add((module, key_name, "*"))
            continue

        for code in [c.strip() for c in langs_str.split(",")]:
            lang = code_to_lang.get(code, "")
            if lang:
                ignore.add((module, key_name, lang))

    return ignore


def _in_ignore(ignore_set: set, module: str, key: str, lang: str) -> bool:
    """Check ignore_set with wildcard support for ALL_LANGUAGES entries."""
    return (module, key, lang) in ignore_set or (module, key, "*") in ignore_set


# ── Scan-report → Ignore-table converter ─────────────────────────────────────

def _detect_lang_sheets(xl: "pd.ExcelFile") -> dict:
    """
    Return {sheet_name: LANG_CODE} for every language-analysis sheet.
    Detects by finding the first 2–5 consecutive uppercase letters in the
    sheet name (e.g. "FIL Extracted", "FIL_Extraction", "HEB_Extraction").
    Sheet must also contain a Comment column to avoid false positives.
    """
    result = {}
    for name in xl.sheet_names:
        m = re.search(r'[A-Z]{2,5}', name)
        if not m:
            continue
        try:
            header = pd.read_excel(xl, sheet_name=name, nrows=0)
        except Exception:
            continue
        cols = [str(c).strip() for c in header.columns]
        if len(cols) >= 6 and "Comment" in cols:
            result[name] = m.group(0)
    return result


def convert_scan_to_ignore(input_xlsx: Path, output_xlsx: Path, log) -> int:
    """
    Convert a language scan report to an ignore table Excel.
    Auto-detects language sheets via column names or sheet name pattern;
    for each Pass row assigns only that sheet's own language.
    Returns number of rows written.
    """
    xl = pd.ExcelFile(str(input_xlsx))

    lang_sheets = _detect_lang_sheets(xl)

    if not lang_sheets:
        raise ValueError(
            "找不到語言分析 Sheet。\n"
            "支援的格式（擇一即可）：\n"
            "  • 欄位名稱含「<語言> Translation」，例如：FIL Translation\n"
            "  • Sheet 名稱以語言代碼開頭，例如：FIL Extracted、FIL_Extraction"
        )

    log(f"🔍 偵測到語言 Sheet: {list(lang_sheets.values())}")

    records: dict[tuple, set] = {}
    for sheet, lang in lang_sheets.items():
        df = pd.read_excel(xl, sheet_name=sheet)
        if df.shape[1] < 6:
            log(f"⚠️  '{sheet}' 欄位不足，跳過")
            continue
        pass_rows = df[df.iloc[:, 5].astype(str).str.strip().str.lower() == "pass"]
        count = 0
        for _, row in pass_rows.iterrows():
            module   = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            enu      = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            keys_raw = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
            for key in keys_raw.split("\n"):
                key = key.strip()
                if key:
                    records.setdefault((module, enu, key), set()).add(lang)
                    count += 1
        log(f"  {lang}: {len(pass_rows)} 列 Pass（{count} 個 Key）")

    rows = []
    for (module, enu, key), langs in sorted(records.items()):
        rows.append([module, enu, key, ", ".join(sorted(langs))])

    wb = Workbook()
    ws = wb.active
    ws.title = "Ignore Table"

    hdr_fill  = PatternFill("solid", fgColor="4472C4")
    hdr_font  = Font(name="Microsoft JhengHei UI", bold=True, color="FFFFFF", size=11)
    thin_s    = Side(style="thin", color="BFBFBF")
    cell_bdr  = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    alt_fill  = PatternFill("solid", fgColor="EEF2FA")
    data_font = Font(name="Microsoft JhengHei UI", size=10)

    for c, h in enumerate(["Module", "ENU翻譯檔字串", "問題Key", "語言"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_bdr

    for r, row_data in enumerate(rows, 2):
        fill = alt_fill if r % 2 == 0 else None
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = data_font
            cell.alignment = Alignment(vertical="center")
            cell.border    = cell_bdr
            if fill:
                cell.fill = fill

    ws.row_dimensions[1].height = 20
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

    wb.save(str(output_xlsx))
    log(f"✅ 輸出完成: {output_xlsx.name}（共 {len(rows)} 筆）")
    return len(rows)


# ── Language scan report (scan mode) ─────────────────────────────────────────

def generate_scan_report(index: dict, output_xlsx: Path, log,
                         scan_langs: list[str],
                         target_langs: dict | None = None,
                         ignore_set: set = None,
                         cancel_event=None,
                         has_module: bool = True):
    """
    Scan ALL strings in the index for selected languages.
    Output: one report sheet grouped by Module.
    """
    lang_label = {v: k for k, v in target_langs.items()} if target_langs else {}

    log(f"🔎 掃描語言: {[lang_label.get(l, l) for l in scan_langs]}")

    groups: dict[tuple, list] = OrderedDict()

    all_records = []
    for en_val, proj_dict in index.items():
        for proj, trans in proj_dict.items():
            all_records.append((proj, en_val, trans))

    total = len(all_records)
    for i, (module, en_val, trans) in enumerate(all_records):
        if cancel_event and cancel_event.is_set():
            log("⚠️  已取消"); return
        if (i + 1) % 200 == 0 or i == total - 1:
            log(f"  掃描第 {i+1} / {total} 筆...")

        enu_val     = trans.get("en", "")
        key_missing = trans.get("_key_missing", {})
        all_keys    = trans.get("_all_keys", [])

        for lang in scan_langs:
            if lang == "en": continue
            val   = trans.get(lang, "")
            label = lang_label.get(lang, lang)
            missing_keys = key_missing.get(lang, [])  # keys specifically missing this lang

            if not val or is_garbled(val):
                issue    = "未翻譯 / 空白"
                keys_str = "\n".join(all_keys)
            elif enu_val and val.strip() == enu_val.strip():
                # Check ignore: pass if ALL keys for this string are in ignore_set
                if ignore_set and all(
                    _in_ignore(ignore_set, module, k, lang) for k in (all_keys or [""])
                ):
                    continue
                issue    = "與英文翻譯檔相同"
                keys_str = "\n".join(all_keys)
            elif missing_keys:
                issue    = "未翻譯 / 空白"
                keys_str = "\n".join(missing_keys)
            else:
                continue

            grp_key = (module, enu_val, issue, keys_str) if has_module else (enu_val, issue, keys_str)
            if grp_key not in groups:
                groups[grp_key] = []
            if label not in groups[grp_key]:
                groups[grp_key].append(label)

    if has_module:
        sorted_groups = sorted(groups.items(), key=lambda x: (x[0][0], x[0][2], x[0][1]))
    else:
        sorted_groups = sorted(groups.items(), key=lambda x: (x[0][1], x[0][0]))
    log(f"📋 發現 {len(sorted_groups)} 筆問題")

    ISSUE_COLORS = {
        "未翻譯 / 空白":    "FF6B6B",
        "與英文翻譯檔相同": "FFD740",
    }

    wb = Workbook()
    rpt = wb.active
    rpt.title = "語言掃描報告"
    RPT_H_FILL = PatternFill("solid", fgColor="C0392B")

    _hdrs = (["Module", "ENU 翻譯檔字串", "問題 Key", "語言", "問題類型"]
             if has_module else
             ["ENU 翻譯檔字串", "問題 Key", "語言", "問題類型"])
    for ci, h in enumerate(_hdrs, 1):
        c = rpt.cell(row=1, column=ci, value=h)
        c.fill = RPT_H_FILL
        c.font = Font(bold=True, color="FFFFFF", name="Microsoft JhengHei UI", size=10)
        c.alignment, c.border = CENTER, BORDER
    rpt.row_dimensions[1].height = 24

    _prev_grp = None
    for ri, (grp_key, langs_list) in enumerate(sorted_groups, 2):
        if has_module:
            module, enu_val, issue, keys_str = grp_key
        else:
            enu_val, issue, keys_str = grp_key
            module = ""
        is_new = (grp_key[0] != (_prev_grp[0] if _prev_grp else None))
        _prev_grp = grp_key
        row_border = Border(
            left=thin, right=thin,
            top=Side(style="medium" if is_new else "thin",
                     color="C0392B" if is_new else "DDDDDD"),
            bottom=thin
        )
        row_h = max(15, len(keys_str.split("\n")) * 14) if keys_str else 15
        rpt.row_dimensions[ri].height = row_h

        _row_vals = ([module, enu_val, keys_str, ", ".join(langs_list), issue]
                     if has_module else
                     [enu_val, keys_str, ", ".join(langs_list), issue])
        _issue_ci = 5 if has_module else 4
        _key_ci   = 3 if has_module else 2
        for ci, v in enumerate(_row_vals, 1):
            c = rpt.cell(row=ri, column=ci, value=_xl_safe(v) or None)
            c.border = row_border
            if ci == _issue_ci:
                c.fill = PatternFill("solid", fgColor=ISSUE_COLORS.get(issue, "FFFFFF"))
                c.font = Font(name="Microsoft JhengHei UI", size=10, bold=True)
                c.alignment = LEFT
            elif ci == _key_ci:
                c.font = Font(name="Microsoft JhengHei UI", size=10)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                c.font = Font(name="Microsoft JhengHei UI", size=10)
                c.alignment = LEFT

    if has_module:
        rpt.column_dimensions["A"].width = 40
        rpt.column_dimensions["B"].width = 44
        rpt.column_dimensions["C"].width = 36
        rpt.column_dimensions["D"].width = 32
        rpt.column_dimensions["E"].width = 22
        rpt.freeze_panes = "A2"
        _auto_row_height(rpt, {"A": 40, "B": 44, "C": 36, "D": 32, "E": 22}, content_cols={"B", "C"})
    else:
        rpt.column_dimensions["A"].width = 50
        rpt.column_dimensions["B"].width = 36
        rpt.column_dimensions["C"].width = 32
        rpt.column_dimensions["D"].width = 22
        rpt.freeze_panes = "A2"
        _auto_row_height(rpt, {"A": 50, "B": 36, "C": 32, "D": 22}, content_cols={"A", "B"})

    wb.save(str(output_xlsx))
    log(f"✅ 輸出完成: {output_xlsx.name}")


def compare_zips(old_index: dict, new_index: dict, output_xlsx: Path, log,
                 cancel_event=None, has_module: bool = True):
    """Compare two pre-built indexes; write strings added in new_index vs old_index."""
    old_keys = set(old_index.keys())
    new_keys = set(new_index.keys())
    added    = sorted(new_keys - old_keys)
    log(f"📊 舊版: {len(old_keys)} 個字串 / 新版: {len(new_keys)} 個字串 / 新增: {len(added)} 個")

    rows = []
    for en in added:
        if cancel_event and cancel_event.is_set():
            log("⚠️  已取消"); return
        for module, proj_data in sorted(new_index[en].items()):
            keys = proj_data.get("_all_keys", [])
            for key in (keys or [""]):
                if has_module:
                    rows.append((module, en, key))
                else:
                    rows.append((en, key))

    if has_module:
        rows.sort(key=lambda x: (x[0], x[1], x[2]))
    else:
        rows.sort(key=lambda x: (x[0], x[1]))

    _alt_fill = PatternFill("solid", fgColor="F5F3FF")

    wb = Workbook()
    ws = wb.active
    ws.title = "新增字串"

    _hdrs = ["Module", "New Strings (EN)", "Key"] if has_module else ["New Strings (EN)", "Key"]
    for ci, h in enumerate(_hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER; c.font = H_FONT; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[1].height = 24

    for ri, row_vals in enumerate(rows, 2):
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = D_FONT; c.alignment = LEFT; c.border = BORDER
            if ri % 2 == 0:
                c.fill = _alt_fill

    n_cols = len(_hdrs)
    last_col = get_column_letter(n_cols)
    if has_module:
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 44
        ws.column_dimensions["C"].width = 44
        col_widths = {"A": 38, "B": 44, "C": 44}
    else:
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 44
        col_widths = {"A": 50, "B": 44}
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{last_col}{len(rows)+1}"
    _auto_row_height(ws, col_widths, content_cols={"A", "B"} if not has_module else {"B", "C"})

    wb.save(str(output_xlsx))
    log(f"✅ 輸出完成: {output_xlsx.name}（{len(rows)} 列，{len(added)} 個新字串）")


def _auto_row_height(ws, col_widths: dict, content_cols: set = None,
                     min_h: int = 15, line_h: int = 15):
    """
    Set row heights based on cell content and column width.
    content_cols: set of column letters to include in calculation.
                  If None, use all columns.
    Uses a 0.7 correction factor for proportional fonts.
    """
    from math import ceil
    FONT_FACTOR = 0.7   # proportional font avg char width vs Excel width unit
    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            if not cell.value: continue
            col_letter = get_column_letter(cell.column)
            if content_cols and col_letter not in content_cols:
                continue
            col_w   = col_widths.get(col_letter, 20)
            eff_w   = col_w / FONT_FACTOR   # effective character capacity
            text    = str(cell.value)
            lines   = 0
            for part in text.split("\n"):
                lines += max(1, ceil(len(part) / max(eff_w, 1)))
            max_lines = max(max_lines, lines)
        ws.row_dimensions[row[0].row].height = max(min_h, max_lines * line_h)


# ── Excel sub-builders ───────────────────────────────────────────────────────

def _build_test_sheet(wb, test_rows_dedup: list, log, has_module: bool = True):
    """
    Append Test Sheet to wb.
    test_rows_dedup: list of (lang_str, page, module, en, trans_or_None, fill_hex)
    has_module: if False, skip the Module column (Web format).
    """
    _lang_counts  = Counter(_tr[0] for _tr in test_rows_dedup)
    _n_langs      = len(_lang_counts)
    _MAX_T        = 6
    _n_testers    = max(2, min(_MAX_T, max(1, _n_langs // 2)))
    _langs_by_cnt = [_l for _l, _ in sorted(_lang_counts.items(), key=lambda x: -x[1])]

    def _greedy(n):
        _h = [(0, i + 1) for i in range(min(n, _n_langs))]
        heapq.heapify(_h)
        _res = {}
        for _l in _langs_by_cnt:
            _tot, _ti = heapq.heappop(_h)
            _res[_l] = _ti
            heapq.heappush(_h, (_tot + _lang_counts[_l], _ti))
        return _res
    _precomp = {_n: _greedy(_n) for _n in range(2, _MAX_T + 1)}

    _col_off = 0 if has_module else -1
    _CFG_COL = 12 + _col_off; _NAME_COL = 13 + _col_off
    _ASGN_LANG = 15 + _col_off; _ASGN_START = 16 + _col_off

    ts = wb.create_sheet("Test Sheet")
    TS_TR_FILL   = PatternFill("solid", fgColor="D9EAD3")
    TS_TR_FONT   = Font(bold=True, color="000000", name="Microsoft JhengHei UI", size=10)
    TS_EDIT_FILL = PatternFill("solid", fgColor="FFF9C4")
    TS_CFG_FILL  = PatternFill("solid", fgColor="302B63")
    TS_SUM_FILL  = PatternFill("solid", fgColor="F5F5F5")
    TS_TTL_FILL  = PatternFill("solid", fgColor="4A4580")
    TS_SRV_FILL  = PatternFill("solid", fgColor="FFF3CD")
    _GRP_FILLS   = [PatternFill("solid", fgColor="FFE3F2FD"),
                    PatternFill("solid", fgColor="FFFFFFFF")]

    _name_rng   = f"$M$3:$M${2 + _MAX_T}"
    _lang_rng   = f"$O$2:$O${1 + _n_langs}"
    _data_rng   = f"$P$2:$T${1 + _n_langs}"
    # Custom override col (U): manager picks a tester name from dropdown;
    # takes priority over greedy auto-assignment when non-empty.
    _CUSTOM_COL = _ASGN_START + (_MAX_T - 1)        # col U (= 21)
    _cust_col_l = get_column_letter(_CUSTOM_COL)     # "U"
    _custom_rng = f"${_cust_col_l}$2:${_cust_col_l}${1 + _n_langs}"

    _data_hdrs = (["Language", "Page", "Module", "EN (Base)", "Translation", "Tester", "Test result"]
                  if has_module else
                  ["Language", "Page", "EN (Base)", "Translation", "Tester", "Test result"])
    for _ci, _h in enumerate(_data_hdrs, 1):
        _c = ts.cell(row=1, column=_ci, value=_h)
        _c.alignment = CENTER; _c.border = BORDER
        if _h in ("Tester", "Test result"):
            _c.fill = TS_TR_FILL; _c.font = TS_TR_FONT
        else:
            _c.fill = HEADER; _c.font = H_FONT
    ts.row_dimensions[1].height = 32

    _n_data = len(test_rows_dedup)
    _prev_lang = None; _grp_idx = -1

    for _r, (_lang, _page, _module, _en, _trans, _fill_hex) in enumerate(test_rows_dedup, 2):
        if _lang != _prev_lang:
            _grp_idx += 1; _prev_lang = _lang
        _row_bg    = _GRP_FILLS[_grp_idx % 2]
        _is_server = (_trans is None)

        _LANG_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
        _row_vals = ([_lang, _page, _module, _en, _trans, None, None]
                     if has_module else
                     [_lang, _page, _en, _trans, None, None])
        for _ci, _v in enumerate(_row_vals, 1):
            _c = ts.cell(row=_r, column=_ci, value=_v)
            _c.font = Font(name="Microsoft JhengHei UI", size=10)
            _c.alignment = _LANG_ALIGN if _ci == 1 else LEFT
            _c.border = BORDER
            if _ci == 5:
                _c.fill = TS_SRV_FILL if _is_server else PatternFill("solid", fgColor="FF" + _fill_hex)
            else:
                _c.fill = _row_bg

        _TESTER_COL = 6 if has_module else 5
        _tc = ts.cell(row=_r, column=_TESTER_COL)
        # Priority: custom name (col U/T) → greedy auto-assignment
        _tc.value = (
            f'=IFERROR('
            f'IF(INDEX({_custom_rng},MATCH(A{_r},{_lang_rng},0))<>"",'
            f'INDEX({_custom_rng},MATCH(A{_r},{_lang_rng},0)),'
            f'INDEX({_name_rng},INDEX({_data_rng},MATCH(A{_r},{_lang_rng},0),MAX(1,MIN(5,$M$2-1)))'
            f')),"")'
        )
        _tc.font = Font(name="Microsoft JhengHei UI", size=10)
        _tc.alignment = CENTER; _tc.border = BORDER; _tc.fill = _row_bg
        ts.row_dimensions[_r].height = 18

    if has_module:
        for _col, _w in zip("ABCDEFG", [18, 10, 36, 42, 36, 18, 20]):
            ts.column_dimensions[_col].width = _w
        ts.freeze_panes = "A2"
        ts.auto_filter.ref = f"A1:G{_n_data + 1}"
    else:
        for _col, _w in zip("ABCDEF", [18, 10, 42, 36, 18, 20]):
            ts.column_dimensions[_col].width = _w
        ts.freeze_panes = "A2"
        ts.auto_filter.ref = f"A1:F{_n_data + 1}"

    _STAT_LANG_COL = 9 + _col_off
    _STAT_CNT_COL  = 10 + _col_off
    for _col, _h in zip([_STAT_LANG_COL, _STAT_CNT_COL], ["Language", "# Strings"]):
        _c = ts.cell(row=1, column=_col, value=_h)
        _c.fill = HEADER; _c.font = H_FONT; _c.alignment = CENTER
    for _i, (_l, _cnt) in enumerate(sorted(_lang_counts.items()), 2):
        _lc = ts.cell(row=_i, column=_STAT_LANG_COL, value=_l)
        _lc.font = Font(name="Microsoft JhengHei UI", size=10); _lc.fill = TS_SUM_FILL; _lc.alignment = LEFT
        _nc = ts.cell(row=_i, column=_STAT_CNT_COL, value=_cnt)
        _nc.font = Font(name="Microsoft JhengHei UI", size=10, bold=True); _nc.fill = TS_SUM_FILL; _nc.alignment = CENTER
    _sum_ttl_r = 2 + _n_langs
    for _col, _v in [(_STAT_LANG_COL, "Total"), (_STAT_CNT_COL, _n_data)]:
        _tc2 = ts.cell(row=_sum_ttl_r, column=_col, value=_v)
        _tc2.font = Font(name="Microsoft JhengHei UI", size=10, bold=True, color="FFFFFF")
        _tc2.fill = TS_TTL_FILL; _tc2.alignment = CENTER

    _cfg_hdr = ts.cell(row=1, column=_CFG_COL, value="Tester 設定")
    _cfg_hdr.fill = TS_CFG_FILL; _cfg_hdr.font = H_FONT; _cfg_hdr.alignment = CENTER
    ts.merge_cells(start_row=1, start_column=_CFG_COL, end_row=1, end_column=_NAME_COL)

    _lbl_font = Font(name="Microsoft JhengHei UI", size=10)
    ts.cell(row=2, column=_CFG_COL, value="人數 / # Testers").font = _lbl_font
    _cnt_v = ts.cell(row=2, column=_NAME_COL, value=_n_testers)
    _cnt_v.font = Font(name="Microsoft JhengHei UI", size=10, bold=True)
    _cnt_v.fill = TS_EDIT_FILL; _cnt_v.alignment = CENTER

    for _ti in range(1, _MAX_T + 1):
        _rr = 2 + _ti
        ts.cell(row=_rr, column=_CFG_COL, value=f"Tester {_ti}").font = _lbl_font
        _nm = ts.cell(row=_rr, column=_NAME_COL, value=f"Tester {_ti}")
        _nm.fill = TS_EDIT_FILL
        _nm.font = Font(name="Microsoft JhengHei UI", size=10, bold=True); _nm.alignment = CENTER

    _note_r = 3 + _MAX_T
    _note = ts.cell(row=_note_r, column=_CFG_COL,
                    value="改 M2（2–6）或姓名可自動更新 ｜ 自定欄直接選名字可覆蓋自動分配")
    _note.font = Font(name="Microsoft JhengHei UI", size=8, color="888888", italic=True)
    ts.merge_cells(start_row=_note_r, start_column=_CFG_COL,
                   end_row=_note_r,   end_column=_NAME_COL)

    _ah_font = Font(name="Microsoft JhengHei UI", size=9, bold=True, color="FFFFFF")
    ts.cell(row=1, column=_ASGN_LANG, value="Language").fill = TS_CFG_FILL
    ts.cell(row=1, column=_ASGN_LANG).font      = _ah_font
    ts.cell(row=1, column=_ASGN_LANG).alignment = CENTER
    for _ni, _n in enumerate(range(2, _MAX_T + 1)):
        _hc = ts.cell(row=1, column=_ASGN_START + _ni, value=f"n={_n}")
        _hc.fill = TS_CFG_FILL; _hc.font = _ah_font; _hc.alignment = CENTER

    # Custom override column header
    _ch = ts.cell(row=1, column=_CUSTOM_COL, value="自定")
    _ch.fill = TS_EDIT_FILL
    _ch.font = Font(name="Microsoft JhengHei UI", size=9, bold=True); _ch.alignment = CENTER

    # Dropdown validation: choices = tester names from M3:M8
    _dv = DataValidation(type="list", formula1=_name_rng,
                         allow_blank=True, showDropDown=False)
    ts.add_data_validation(_dv)

    for _ai, _l in enumerate(_langs_by_cnt, 2):
        _lc = ts.cell(row=_ai, column=_ASGN_LANG, value=_l)
        _lc.font = Font(name="Microsoft JhengHei UI", size=9); _lc.fill = TS_SUM_FILL; _lc.alignment = LEFT
        for _ni, _n in enumerate(range(2, _MAX_T + 1)):
            _vc = ts.cell(row=_ai, column=_ASGN_START + _ni, value=_precomp[_n].get(_l, 0))
            _vc.font = Font(name="Microsoft JhengHei UI", size=9, bold=True)
            _vc.fill = TS_SUM_FILL; _vc.alignment = CENTER
        # Custom col: editable cell with dropdown
        _cc = ts.cell(row=_ai, column=_CUSTOM_COL, value="")
        _cc.font = Font(name="Microsoft JhengHei UI", size=9)
        _cc.fill = TS_EDIT_FILL; _cc.alignment = CENTER
        _dv.add(_cc)

    ts.column_dimensions[get_column_letter(8 + _col_off)].width  = 4   # spacer before stats
    ts.column_dimensions[get_column_letter(_STAT_LANG_COL)].width = 22
    ts.column_dimensions[get_column_letter(_STAT_CNT_COL)].width  = 12
    ts.column_dimensions[get_column_letter(11 + _col_off)].width  = 4   # spacer before config
    ts.column_dimensions[get_column_letter(_CFG_COL)].width  = 22
    ts.column_dimensions[get_column_letter(_NAME_COL)].width = 18
    ts.column_dimensions[get_column_letter(14 + _col_off)].width  = 4   # spacer before assignment
    ts.column_dimensions[get_column_letter(_ASGN_LANG)].width = 22
    for _ni in range(_MAX_T - 1):
        ts.column_dimensions[get_column_letter(_ASGN_START + _ni)].width = 7
    ts.column_dimensions[_cust_col_l].width = 14   # custom col

    log(f"🧪 Test Sheet: {_n_data} 筆（{_n_langs} 個語言，預設 {_n_testers} 位 Tester，支援 2–{_MAX_T} 人）")


def _build_issue_sheet(wb, rows_data: list, sorted_langs: list,
                       lang_label: dict, ignore_set, has_module: bool = True):
    """Append 翻譯問題報告 sheet to wb. Returns (n_groups, n_issues)."""
    rpt = wb.create_sheet("翻譯問題報告")
    RPT_H_FILL = PatternFill("solid", fgColor="C0392B")

    _issue_hdrs = (["Module", "ENU 翻譯檔字串", "問題 Key", "語言", "問題類型"]
                   if has_module else
                   ["ENU 翻譯檔字串", "問題 Key", "語言", "問題類型"])
    for ci, h in enumerate(_issue_hdrs, 1):
        c = rpt.cell(row=1, column=ci, value=h)
        c.fill = RPT_H_FILL
        c.font = Font(bold=True, color="FFFFFF", name="Microsoft JhengHei UI", size=10)
        c.alignment, c.border = CENTER, BORDER
    rpt.row_dimensions[1].height = 24

    ISSUE_COLORS = {"未翻譯 / 空白": "FF6B6B", "與英文翻譯檔相同": "FFD740"}
    rpt_groups: dict[tuple, list] = OrderedDict()

    for rd in rows_data:
        if rd["not_found"]: continue
        trans             = rd["trans"]
        module            = rd["module"]
        enu_val           = rd.get("enu_val", "")
        key_name          = rd.get("key_name", "")
        key_missing_langs = rd.get("key_missing_langs", [])

        for lang in sorted_langs:
            if lang == "en": continue
            val   = trans.get(lang, "")
            label = lang_label.get(lang, lang)

            if not val or is_garbled(val):
                issue = "未翻譯 / 空白"; keys_str = key_name
            elif enu_val and val.strip() == enu_val.strip():
                if ignore_set and _in_ignore(ignore_set, module, key_name, lang):
                    continue
                issue = "與英文翻譯檔相同"; keys_str = key_name
            elif lang in key_missing_langs:
                issue = "未翻譯 / 空白"; keys_str = key_name
            else:
                continue

            grp_key = ((module, enu_val, issue, keys_str) if has_module
                       else (enu_val, issue, keys_str))
            if grp_key not in rpt_groups:
                rpt_groups[grp_key] = []
            if label not in rpt_groups[grp_key]:
                rpt_groups[grp_key].append(label)

    if has_module:
        sorted_groups = sorted(rpt_groups.items(), key=lambda x: (x[0][0], x[0][2], x[0][1]))
    else:
        sorted_groups = sorted(rpt_groups.items(), key=lambda x: (x[0][1], x[0][0]))

    _prev_grp = None
    for ri, (grp_key, langs_list) in enumerate(sorted_groups, 2):
        if has_module:
            module, enu_val, issue, keys_str = grp_key
        else:
            enu_val, issue, keys_str = grp_key
            module = ""
        is_new_grp = (grp_key[0] != (_prev_grp[0] if _prev_grp else None))
        _prev_grp = grp_key
        row_border = Border(
            left=thin, right=thin,
            top=Side(style="medium" if is_new_grp else "thin",
                     color="C0392B" if is_new_grp else "DDDDDD"),
            bottom=thin)
        row_h = max(15, len(keys_str.split("\n")) * 14) if keys_str else 15
        rpt.row_dimensions[ri].height = row_h

        _row_vals = ([module, enu_val, keys_str, ", ".join(langs_list), issue]
                     if has_module else
                     [enu_val, keys_str, ", ".join(langs_list), issue])
        _issue_ci = 5 if has_module else 4
        _key_ci   = 3 if has_module else 2
        for ci, v in enumerate(_row_vals, 1):
            c = rpt.cell(row=ri, column=ci, value=v or None)
            c.border = row_border
            if ci == _issue_ci:
                c.fill = PatternFill("solid", fgColor=ISSUE_COLORS.get(issue, "FFFFFF"))
                c.font = Font(name="Microsoft JhengHei UI", size=10, bold=True)
                c.alignment = LEFT
            elif ci == _key_ci:
                c.font = Font(name="Microsoft JhengHei UI", size=10)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                c.font = Font(name="Microsoft JhengHei UI", size=10)
                c.alignment = LEFT

    if has_module:
        rpt.column_dimensions["A"].width = 40
        rpt.column_dimensions["B"].width = 44
        rpt.column_dimensions["C"].width = 36
        rpt.column_dimensions["D"].width = 32
        rpt.column_dimensions["E"].width = 22
        rpt.freeze_panes = "A2"
        _auto_row_height(rpt, {"A": 40, "B": 44, "C": 36, "D": 32, "E": 22}, content_cols={"B", "C"})
    else:
        rpt.column_dimensions["A"].width = 50
        rpt.column_dimensions["B"].width = 36
        rpt.column_dimensions["C"].width = 32
        rpt.column_dimensions["D"].width = 22
        rpt.freeze_panes = "A2"
        _auto_row_height(rpt, {"A": 50, "B": 36, "C": 32, "D": 22}, content_cols={"A", "B"})
    return len(sorted_groups), sum(len(v) for v in rpt_groups.values())


def _build_legend_sheet(wb):
    """Append 說明 sheet (colour legend + lang code table) to wb."""
    ls = wb.create_sheet("說明")
    for ri, (a, b) in enumerate([
        ("色彩說明", ""),
        ("  黃色底", "該語言在此 module 翻譯中字串最長（唯一最長才標示）"),
        ("  紅色底", "未翻譯、空白、亂碼、或翻譯與英文相同"),
        ("  橘色字", "模糊比對（Excel 數字對應翻譯檔變數，hover 看原始字串）"),
        ("  黃色警示欄", "查無字串（此英文字串在翻譯檔中找不到）"),
        ("", ""),
        ("翻譯問題報告色彩", ""),
        ("  紅底", "未翻譯 / 空白"),
        ("  黃底", "與英文翻譯檔相同（ENU）"),
        ("", ""), ("語言代碼", "語言名稱"),
    ] + list(LANG_NAMES.items()), 1):
        ca = ls.cell(row=ri, column=1, value=a)
        cb = ls.cell(row=ri, column=2, value=b)
        bold = ri in (1, 7, 12)
        ca.font = Font(name="Microsoft JhengHei UI", size=10, bold=bold)
        cb.font = Font(name="Microsoft JhengHei UI", size=10, bold=bold)
        if ri == 2:  ca.fill = YELLOW
        if ri == 3:  ca.fill = RED
        if ri == 8:  ca.fill = PatternFill("solid", fgColor="FF6B6B")
        if ri == 9:  ca.fill = PatternFill("solid", fgColor="FFAB40")
        if ri == 10: ca.fill = PatternFill("solid", fgColor="FFD740")
    ls.column_dimensions["A"].width = 22
    ls.column_dimensions["B"].width = 55


# ── Excel generator ───────────────────────────────────────────────────────────

def generate_excel(index: dict, norm_index: dict, input_xlsx: Path,
                   output_xlsx: Path, log, target_langs: dict | None = None,
                   ignore_set: set = None, cancel_event=None,
                   has_module: bool = True):
    import pandas as pd_
    import fnmatch as _fnmatch
    df       = pd_.read_excel(str(input_xlsx), header=0)
    page_col = df.columns[0]
    en_col   = df.columns[1]

    # Pre-expand wildcard rows (* glob) into individual query tuples
    query_rows = []
    for _, row in df.iterrows():
        en_raw = str(row[en_col]).strip()   if pd_.notna(row[en_col])   else ""
        pg     = str(row[page_col]).strip() if pd_.notna(row[page_col]) else ""
        if '*' in en_raw:
            hits = sorted(k for k in index if _fnmatch.fnmatch(k.lower(), en_raw.lower()))
            if hits:
                query_rows.extend((pg, m) for m in hits)
            else:
                query_rows.append((pg, en_raw))   # keep to show ⚠ not-found
        else:
            query_rows.append((pg, en_raw))
    total = len(query_rows)

    rows_data = []
    seen_langs: set = set()

    for i, (page, en) in enumerate(query_rows):
        if cancel_event and cancel_event.is_set():
            log("⚠️  已取消"); return

        if (i + 1) % 5 == 0 or i == total - 1:
            log(f"  處理第 {i+1} / {total} 筆...")

        rec = index.get(en)
        if not rec and en:
            for k, v in index.items():
                if k.lower() == en.lower(): rec = v; break

        matched_key = en
        if not rec and en:
            nq = _normalize(en)
            orig = norm_index.get(nq)
            if orig:
                rec = index.get(orig)
                matched_key = orig

        if rec:
            first_proj = sorted(rec.keys())[0]
            for proj, proj_trans in sorted(rec.items()):
                all_keys  = proj_trans.get("_all_keys", [])
                key_trans = proj_trans.get("_key_trans", {})
                seen_langs.update(k for k in proj_trans if not k.startswith("_"))

                if len(all_keys) > 1:
                    # Multiple keys → one row per key
                    key_missing = proj_trans.get("_key_missing", {})
                    for ki, key in enumerate(all_keys):
                        kt = key_trans.get(key, {})
                        is_first_row = (proj == first_proj and ki == 0)
                        # For this key: which langs are missing (only if this key is in key_missing)
                        this_key_missing_langs = [l for l, keys in key_missing.items() if key in keys]
                        rows_data.append({
                            "page":            page,
                            "en":              en,
                            "module":          proj,
                            "key_name":        key,
                            "trans":           kt,
                            "enu_val":         proj_trans.get("en", ""),
                            "is_first":        is_first_row,
                            "fuzzy":           matched_key != en,
                            "fuzzy_key":       matched_key,
                            "not_found":       False,
                            "key_missing_langs": this_key_missing_langs,  # langs missing for THIS key
                        })
                else:
                    # Single key → one row as before
                    rows_data.append({
                        "page":            page,
                        "en":              en,
                        "module":          proj,
                        "key_name":        all_keys[0] if all_keys else "",
                        "trans":           {k: v for k, v in proj_trans.items() if not k.startswith("_")},
                        "enu_val":         proj_trans.get("en", ""),
                        "is_first":        proj == first_proj,
                        "fuzzy":           matched_key != en,
                        "fuzzy_key":       matched_key,
                        "not_found":       False,
                        "key_missing_langs": [],
                    })
        else:
            rows_data.append({
                "page": page, "en": en, "module": "",
                "key_name": "", "trans": {}, "enu_val": "",
                "is_first": True, "fuzzy": False,
                "fuzzy_key": en, "not_found": True,
            })

    if target_langs:
        allowed   = set(target_langs.values())
        key_order = {v: i for i, v in enumerate(target_langs.values())}
        sorted_langs = sorted(
            [l for l in seen_langs if l in allowed and not _is_regional_variant(l)],
            key=lambda l: key_order.get(l, 999)
        )
        lang_label = {v: k for k, v in target_langs.items()}
    else:
        filtered     = [l for l in seen_langs if not _is_regional_variant(l)]
        sorted_langs = ([l for l in LANG_PRIORITY if l in filtered] +
                        sorted(l for l in filtered if l not in LANG_PRIORITY))
        lang_label   = {}

    log(f"📊 語言: {[lang_label.get(l, l) for l in sorted_langs]}")

    wb = Workbook()

    # ── Sheet 1: 翻譯對照 ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "翻譯對照"

    def _hdr(l):
        code = lang_label.get(l, l)
        name = LANG_NAMES.get(l, "")
        return f"{code}\n{name}" if name else code

    def _ts_lang(l):
        """Single-line language label for Test Sheet: 'ARA (阿拉伯語)'"""
        code = lang_label.get(l, l)
        name = LANG_NAMES.get(l, "")
        return f"{code} ({name})" if name else code

    MODULE_FILL   = PatternFill("solid", fgColor="FFFFFF")
    MODULE_FONT   = Font(name="Microsoft JhengHei UI", size=9, color="302B63")
    NF_FILL       = PatternFill("solid", fgColor="FFF3CD")
    NF_FONT       = Font(name="Microsoft JhengHei UI", size=9, color="cc4400", bold=True)
    DIM_FONT      = Font(name="Microsoft JhengHei UI", size=10, color="aaaaaa")
    GROUP_TOP     = Border(left=thin, right=thin,
                           top=Side(style="medium", color="302B63"),
                           bottom=Side(style="thin", color="DDDDDD"))
    KEY_COL_FILL  = PatternFill("solid", fgColor="F5F3FF")

    _base_hdrs = (["Page", "Module", "EN (Base, from App)", "Key(s)"]
                  if has_module else
                  ["Page", "EN (Base, from App)", "Key(s)"])
    for ci, h in enumerate(_base_hdrs + [_hdr(l) for l in sorted_langs], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = HEADER, H_FONT, CENTER, BORDER
    ws.row_dimensions[1].height = 32
    # Column indices for data rows
    _mod_ci  = 2                          # Module col (only used when has_module)
    _en_ci   = 3 if has_module else 2     # EN col
    _key_ci  = 4 if has_module else 3     # Key col
    _lang_ci = 5 if has_module else 4     # First lang col

    test_rows: list = []   # collects (lang_header, page, module, en, trans, fill_hex) for Test Sheet

    for ri, rd in enumerate(rows_data, 2):
        en_val    = rd["en"]
        trans     = rd["trans"]
        enu_val   = rd.get("enu_val", "")
        key_name  = rd.get("key_name", "")
        is_first  = rd["is_first"]
        not_found = rd["not_found"]
        border    = GROUP_TOP if is_first else BORDER

        valid    = {l: _visual_len(v) for l, v in trans.items()
                    if l in sorted_langs and v and not is_garbled(v)
                    and l != "en" and v.strip() != en_val.strip()}
        max_len  = max(valid.values()) if valid else 0
        max_langs = {l for l, n in valid.items() if n == max_len}

        # If ARA is among the longest, also find the second longest tier
        second_langs: set = set()
        if "ar" in max_langs and len(valid) > len(max_langs):
            second_len   = max((n for l, n in valid.items() if l not in max_langs), default=0)
            second_langs = {l for l, n in valid.items() if n == second_len and l not in max_langs}

        # Col 1: Page — dim on repeated rows within same group
        c = ws.cell(row=ri, column=1, value=rd["page"] if is_first else None)
        c.font = B_FONT if is_first else DIM_FONT
        c.alignment, c.border = LEFT, border

        # Col 2: Module — only when has_module
        if has_module:
            if not_found:
                c = ws.cell(row=ri, column=_mod_ci, value="⚠️ 查無字串")
                c.fill, c.font, c.alignment, c.border = NF_FILL, NF_FONT, LEFT, border
            else:
                c = ws.cell(row=ri, column=_mod_ci, value=rd["module"] or None)
                c.fill = MODULE_FILL
                c.font = MODULE_FONT
                c.alignment, c.border = LEFT, border
        else:
            # No module col — show ⚠ in EN col when not found
            pass

        # EN Base — dim on repeated rows
        c = ws.cell(row=ri, column=_en_ci, value=_xl_safe(en_val) if is_first else None)
        if not has_module and not_found and is_first:
            c.fill, c.font, c.alignment, c.border = NF_FILL, NF_FONT, LEFT, border
        if rd.get("fuzzy") and is_first:
            c.font = Font(name="Microsoft JhengHei UI", size=10, bold=True, color="E65100")
            from openpyxl.comments import Comment
            fuzzy_key = rd["fuzzy_key"]
            note  = f"⚠️ 模糊比對（變數/換行符號已正規化）\n翻譯檔實際字串:\n{fuzzy_key}"
            lines = note.split("\n")
            w = max(280, max(len(l) for l in lines) * 7 + 20)
            h = 20 + len(lines) * 18
            c.comment = Comment(note, "MUI Tool", height=h, width=w)
        else:
            c.font = B_FONT if is_first else DIM_FONT
        c.alignment, c.border = LEFT, border

        # Key col
        c = ws.cell(row=ri, column=_key_ci, value=_xl_safe(key_name) or None)
        c.fill   = KEY_COL_FILL
        c.font = Font(name="Microsoft JhengHei UI", size=10)
        c.alignment = LEFT
        c.border = border

        # Lang cols
        for ci, lang in enumerate(sorted_langs, _lang_ci):
            val     = trans.get(lang, "")
            is_base = (lang == "en")
            if is_base:
                val = enu_val
            c = ws.cell(row=ri, column=ci, value=_xl_safe(val) or None)
            c.alignment, c.border = LEFT, border
            if not val or is_garbled(val):
                c.fill, c.font = RED, D_FONT
            elif not is_base and (
                val.strip() == en_val.strip()
                or (enu_val and val.strip() == enu_val.strip())
            ):
                module = rd.get("module", "")
                if ignore_set and _in_ignore(ignore_set, module, key_name, lang):
                    c.font = D_FONT
                else:
                    c.fill, c.font = RED, D_FONT
            elif lang in max_langs:
                c.fill, c.font = YELLOW, B_FONT
                test_rows.append((_ts_lang(lang), rd["page"], rd["module"], en_val, _xl_safe(val) or val, "FFD700"))
            elif lang in second_langs:
                c.fill, c.font = YELLOW2, B_FONT
                test_rows.append((_ts_lang(lang), rd["page"], rd["module"], en_val, _xl_safe(val) or val, "FFF176"))
            else:
                c.font = D_FONT

        # Not-found strings come from the server — expand one row per language
        if rd["not_found"] and rd["is_first"]:
            for _sl in sorted_langs:
                if _sl == "en": continue
                test_rows.append((_ts_lang(_sl), rd["page"], "⚠️ Server string", en_val, None, "FFF3CD"))

    ws.column_dimensions["A"].width = 16
    if has_module:
        ws.column_dimensions["B"].width = 40   # Module
        ws.column_dimensions["C"].width = 42   # EN
        ws.column_dimensions["D"].width = 36   # Key
    else:
        ws.column_dimensions["B"].width = 42   # EN
        ws.column_dimensions["C"].width = 36   # Key
    _freeze = get_column_letter(_lang_ci) + "2"
    for i in range(len(sorted_langs)):
        ws.column_dimensions[get_column_letter(i + _lang_ci)].width = 28
    ws.freeze_panes = _freeze

    main_col_widths = {"A": 16}
    if has_module:
        main_col_widths.update({"B": 40, "C": 42, "D": 36})
    else:
        main_col_widths.update({"B": 42, "C": 36})
    lang_cols = set()
    for i in range(len(sorted_langs)):
        col = get_column_letter(i + _lang_ci)
        main_col_widths[col] = 28
        lang_cols.add(col)
    _en_col_letter = get_column_letter(_en_ci)
    _auto_row_height(ws, main_col_widths, content_cols=lang_cols | {_en_col_letter})

    # ── Sheet 2: Test Sheet ───────────────────────────────────────────────────
    seen_test: set = set()
    test_rows_dedup: list = []
    for _tr in test_rows:
        _key = (_tr[0], str(_tr[3] or "")) if _tr[4] is None else (_tr[0], str(_tr[4] or "").strip())
        if _key not in seen_test:
            seen_test.add(_key)
            test_rows_dedup.append(_tr)
    test_rows_dedup.sort(key=lambda x: x[0])
    _build_test_sheet(wb, test_rows_dedup, log, has_module=has_module)

    # ── Sheet 3: 翻譯問題報告 ────────────────────────────────────────────────
    n_groups, n_issues = _build_issue_sheet(wb, rows_data, sorted_langs, lang_label, ignore_set,
                                            has_module=has_module)
    log(f"📋 翻譯問題報告: {n_groups} 筆（{n_issues} 個問題）")

    # ── Sheet 4: 說明 ─────────────────────────────────────────────────────────
    _build_legend_sheet(wb)

    wb.save(str(output_xlsx))
    log(f"✅ 輸出完成: {output_xlsx.name}")

# ── GUI ───────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"MUI Translation Tool  {APP_VERSION}")
        self.resizable(True, True)
        self.minsize(700, 520)
        self.configure(bg="#1e1e2e")
        self._cfg = load_config()
        self._zip_path = self._excel_path = self._out_path = None
        self._ignore_path = None
        self._zip_old_path = None   # diff mode: old version zip
        self._diff_out_path = None  # diff mode: output xlsx
        self._scan_out_path = None
        self._convert_in_path = self._convert_out_path = None
        self._cancel_event = threading.Event()
        self._ignore_var = tk.StringVar(value="未設定（可選）")
        self._web_target_langs: dict = {}   # populated when Web zip is loaded

        # Scrollable main canvas
        self._canvas = tk.Canvas(self, bg="#1e1e2e", highlightthickness=0)
        self._vscroll = ttk.Scrollbar(self, orient="vertical",
                                      command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=self._vscroll.set)
        self._vscroll.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        self._inner = tk.Frame(self._canvas, bg="#1e1e2e")
        self._inner_id = self._canvas.create_window(
            (0, 0), window=self._inner, anchor="nw")
        self._inner.columnconfigure(0, weight=1)

        self._inner.bind("<Configure>", self._on_frame_configure)
        self._canvas.bind("<Configure>", self._on_canvas_configure)
        # Mouse wheel scrolling
        self.bind_all("<MouseWheel>",
                      lambda e: self._canvas.yview_scroll(-1*(e.delta//120), "units"))
        self.bind_all("<Button-4>",
                      lambda e: self._canvas.yview_scroll(-1, "units"))
        self.bind_all("<Button-5>",
                      lambda e: self._canvas.yview_scroll(1, "units"))

        self._build_ui()
        self._restore_paths()
        self._center()
        threading.Thread(target=self._check_for_update, daemon=True).start()

    def _on_frame_configure(self, event=None):
        canvas_h = self._canvas.winfo_height()
        req_h = self._inner.winfo_reqheight()
        target_h = max(canvas_h if canvas_h > 1 else req_h, req_h)
        self._canvas.itemconfig(self._inner_id, height=target_h)
        self._canvas.configure(scrollregion=(0, 0, self._canvas.winfo_width(), target_h))

    def _on_canvas_configure(self, event):
        self._canvas.itemconfig(self._inner_id, width=event.width)
        req_h = self._inner.winfo_reqheight()
        target_h = max(event.height, req_h)
        self._canvas.itemconfig(self._inner_id, height=target_h)
        self._canvas.configure(scrollregion=(0, 0, event.width, target_h))

    def _restore_paths(self):
        """Restore last app and its paths from config."""
        last_app = self._cfg.get("last_app")
        if last_app and last_app in APP_CONFIGS:
            self._select_app(last_app, init=True)   # restore app selection (already built UI)
            # Re-highlight correct button
            for name, btn in self._app_buttons.items():
                btn.configure(
                    bg="#cba6f7" if name == last_app else "#313244",
                    fg="#1e1e2e"  if name == last_app else "#a6adc8"
                )
            self._app_var.set(last_app)
            if last_app == "Web":
                self._lang_preview.configure(text="（選取 Zip 後自動偵測語言）")
            else:
                cfg = APP_CONFIGS.get(last_app, {})
                self._lang_preview.configure(text="  ".join(cfg.keys()) if cfg else "（全部語言）")

        app = self._app_var.get()
        app_cfg = self._cfg.get(app, {})

        def _try_set(path_str, var, attr):
            if path_str and Path(path_str).exists():
                setattr(self, attr, Path(path_str))
                var.set(self._fmt_name(Path(path_str).name))

        _try_set(app_cfg.get("zip"),    self._zip_var,    "_zip_path")
        _try_set(app_cfg.get("excel"),  self._excel_var,  "_excel_path")
        _try_set(app_cfg.get("out"),    self._out_var,    "_out_path")
        _try_set(app_cfg.get("ignore"), self._ignore_var, "_ignore_path")

        scan_out = app_cfg.get("scan_out")
        if scan_out and Path(scan_out).parent.exists():
            self._scan_out_path = Path(scan_out)
            self._scan_out_var.set(self._fmt_name(Path(scan_out).name))

        _try_set(app_cfg.get("convert_in"),  self._convert_in_var,  "_convert_in_path")
        _try_set(app_cfg.get("convert_out"), self._convert_out_var, "_convert_out_path")
        _try_set(app_cfg.get("zip_old"),     self._zip_old_var,     "_zip_old_path")
        _try_set(app_cfg.get("diff_out"),    self._diff_out_var,    "_diff_out_path")

        if self._zip_path or self._excel_path:
            self._log(f"📂 已載入上次 {app} 設定")

        # For Web app: detect langs from saved zip so scan checkboxes are populated
        if app == "Web" and self._zip_path:
            self._web_target_langs = _detect_web_langs_from_zip(self._zip_path)
            self._update_web_lang_preview()

        # Rebuild scan checkboxes (also restores saved lang selections)
        self._rebuild_scan_checkboxes()

    def _save_app_paths(self):
        """Save current app's paths to config."""
        app = self._app_var.get()
        if app not in self._cfg:
            self._cfg[app] = {}
        if self._zip_path:
            self._cfg[app]["zip"]    = str(self._zip_path)
        if self._excel_path:
            self._cfg[app]["excel"]  = str(self._excel_path)
        if self._out_path:
            self._cfg[app]["out"]    = str(self._out_path)
        if self._ignore_path:
            self._cfg[app]["ignore"] = str(self._ignore_path)
        if self._scan_out_path:
            self._cfg[app]["scan_out"] = str(self._scan_out_path)
        if self._convert_in_path:
            self._cfg[app]["convert_in"]  = str(self._convert_in_path)
        if self._convert_out_path:
            self._cfg[app]["convert_out"] = str(self._convert_out_path)
        if self._zip_old_path:
            self._cfg[app]["zip_old"]  = str(self._zip_old_path)
        if self._diff_out_path:
            self._cfg[app]["diff_out"] = str(self._diff_out_path)
        if hasattr(self, '_scan_lang_vars'):
            self._cfg[app]["scan_langs"] = [l for l, v in self._scan_lang_vars.items() if v.get()]
        self._cfg["last_app"] = app
        save_config(self._cfg)

    # ── Layout ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        BG = "#1e1e2e"
        # ── Title bar ─────────────────────────────────────────────────────────
        hdr = tk.Frame(self._inner, bg="#11111b")
        hdr.pack(fill="x")
        title_f = tk.Frame(hdr, bg="#11111b")
        title_f.pack(pady=(14, 2))
        tk.Label(title_f, text="MUI Translation Tool",
                 font=("Microsoft JhengHei UI", 16, "bold"), fg="#ffffff", bg="#11111b").pack(side="left")
        tk.Label(title_f, text=f" {APP_VERSION}",
                 font=("Microsoft JhengHei UI", 11, "bold"), fg="#cba6f7", bg="#11111b").pack(side="left", pady=(4,0))
        sub_f = tk.Frame(hdr, bg="#11111b")
        sub_f.pack(pady=(0, 10))
        tk.Label(sub_f, text=f"by {APP_AUTHOR}",
                 font=("Microsoft JhengHei UI", 10), fg="#6c7086", bg="#11111b").pack(side="left")
        tk.Button(sub_f, text="📋 更新記錄",
                  font=("Microsoft JhengHei UI", 9), fg="#cba6f7", bg="#11111b",
                  activeforeground="#cdd6f4", activebackground="#11111b",
                  relief="flat", cursor="hand2", bd=0,
                  command=self._show_changelog).pack(side="left", padx=(12, 0))

        # ── Two-column body ───────────────────────────────────────────────────
        body = tk.Frame(self._inner, bg=BG)
        body.pack(fill="x")
        body.columnconfigure(0, weight=0, minsize=400)
        body.columnconfigure(1, weight=1, minsize=340)
        body.rowconfigure(0, weight=0)   # top two columns fixed

        # ── LEFT COLUMN ───────────────────────────────────────────────────────
        left = tk.Frame(body, bg=BG)
        left.grid(row=0, column=0, sticky="new", padx=(16, 8), pady=12)

        # App selector (2-row grid, equal-width columns)
        app_hdr = tk.Frame(left, bg=BG)
        app_hdr.pack(fill="x", pady=(0, 3))
        tk.Label(app_hdr, text="選擇 App：", font=("Microsoft JhengHei UI", 11, "bold"),
                 fg="#a6adc8", bg=BG).pack(side="left", anchor="n", padx=(0, 8))
        app_btn_area = tk.Frame(app_hdr, bg=BG)
        app_btn_area.pack(side="left")
        _app_names = list(APP_CONFIGS.keys())
        _per_row   = (len(_app_names) + 1) // 2   # ceil(n/2): balanced split
        for c in range(_per_row):
            app_btn_area.columnconfigure(c, uniform="app_btn", weight=1)
        self._app_var     = tk.StringVar(value=_app_names[0])
        self._app_buttons = {}
        for i, app_name in enumerate(_app_names):
            langs = APP_CONFIGS[app_name]
            count = "動態" if app_name == "Web" else (len(langs) if langs else "全部")
            row, col = divmod(i, _per_row)
            btn = tk.Button(app_btn_area, text=f"{app_name}({count})",
                            font=("Microsoft JhengHei UI", 10, "bold"), relief="flat",
                            cursor="hand2", padx=7, pady=4, bd=0,
                            command=lambda n=app_name: self._select_app(n))
            btn.grid(row=row, column=col, sticky="ew",
                     padx=(0, 4), pady=(0, 3))
            self._app_buttons[app_name] = btn
        # Fixed-height container so lang preview never shifts layout below it
        _preview_frame = tk.Frame(left, bg=BG, height=54)
        _preview_frame.pack(fill="x", pady=(0, 8))
        _preview_frame.pack_propagate(False)
        self._lang_preview = tk.Label(_preview_frame, text="", font=("Microsoft JhengHei UI", 10),
                                      fg="#cba6f7", bg=BG,
                                      wraplength=370, justify="left")
        self._lang_preview.pack(anchor="nw", fill="x")
        self._select_app(list(APP_CONFIGS.keys())[0], init=True)

        # Mode selector
        mode_hdr = tk.Frame(left, bg=BG)
        mode_hdr.pack(fill="x", pady=(0, 10))
        tk.Label(mode_hdr, text="選擇模式：", font=("Microsoft JhengHei UI", 11, "bold"),
                 fg="#a6adc8", bg=BG).pack(side="left", padx=(0, 8))
        mode_btns = tk.Frame(mode_hdr, bg=BG)
        mode_btns.pack(side="left")
        self._mode_var     = tk.StringVar(value="excel")
        self._mode_buttons = {}
        # 2×2 grid: equal-width columns (uniform group), 3 px gaps
        mode_btns.columnconfigure(0, weight=1, uniform="mc")
        mode_btns.columnconfigure(1, weight=1, uniform="mc")
        _mode_grid = {"excel": (0, 0), "scan": (0, 1),
                      "convert": (1, 0), "diff": (1, 1)}
        for val, label in [("excel", "Excel 查詢"), ("scan", "語言全掃描"),
                            ("convert", "轉換 Ignore"), ("diff", "比對新字串")]:
            r, c = _mode_grid[val]
            btn = tk.Button(mode_btns, text=label,
                            font=("Microsoft JhengHei UI", 11, "bold"), relief="flat",
                            cursor="hand2", padx=14, pady=5, bd=0,
                            command=lambda v=val: self._set_mode(v))
            btn.grid(row=r, column=c, sticky="ew",
                     padx=(0, 3) if c == 0 else 0,
                     pady=(0, 3) if r == 0 else 0)
            self._mode_buttons[val] = btn
        self._set_mode("excel", init=True)

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=(0, 8))

        # File pickers container — swapped between standard and convert mode
        fp_container = tk.Frame(left, bg=BG)
        fp_container.pack(fill="x")

        fp = tk.Frame(fp_container, bg=BG)
        fp.pack(fill="x")
        self._std_fp = fp
        self._zip_var   = tk.StringVar(value="尚未選取")
        self._excel_var = tk.StringVar(value="尚未選取")
        self._out_var   = tk.StringVar(value="尚未選取")
        self._ignore_var = tk.StringVar(value="未設定（可選）")
        self._zip_row   = self._make_file_row(fp, "📦 翻譯 Zip 檔:",  self._zip_var,   self._pick_zip,   row=0)
        self._excel_row = self._make_file_row(fp, "📄 輸入 Excel:",   self._excel_var, self._pick_excel, row=1)
        self._out_row   = self._make_file_row(fp, "💾 輸出 Excel:",   self._out_var,   self._pick_out,   row=2, is_save=True)
        # Ignore row with clear button
        self._ignore_lbl = tk.Label(fp, text="🚫 Ignore Excel:", font=("Microsoft JhengHei UI", 11), fg="#cdd6f4",
                 bg=BG, anchor="w")
        self._ignore_lbl.grid(row=3, column=0, sticky="w", pady=4)
        self._ignore_val = tk.Label(fp, textvariable=self._ignore_var, font=("Microsoft JhengHei UI", 9), fg="#a6adc8",
                 bg=BG, anchor="w")
        self._ignore_val.grid(row=3, column=1, sticky="ew", padx=(4, 0))
        ig_btns = tk.Frame(fp, bg=BG)
        ig_btns.grid(row=3, column=2, padx=(4, 0))
        tk.Button(ig_btns, text="選取", font=("Microsoft JhengHei UI", 10),
                  bg="#45475a", fg="white", activebackground="#3d59a1",
                  activeforeground="white", relief="flat", padx=8, pady=2,
                  cursor="hand2", command=self._pick_ignore).pack(side="left", padx=(0, 3))
        tk.Button(ig_btns, text="清除", font=("Microsoft JhengHei UI", 10),
                  bg="#2d2030", fg="#f38ba8", activebackground="#3a2535",
                  activeforeground="white", relief="flat", padx=8, pady=2,
                  cursor="hand2", command=self._clear_ignore).pack(side="left")

        # Tooltips for standard file rows
        self._bind_row_tooltip(self._zip_row,   lambda: str(self._zip_path)    if self._zip_path    else None)
        self._bind_row_tooltip(self._excel_row, lambda: str(self._excel_path)  if self._excel_path  else None)
        self._bind_row_tooltip(self._out_row,   lambda: str(self._out_path)    if self._out_path    else None)
        for w in (self._ignore_lbl, self._ignore_val):
            self._bind_tooltip(w, lambda: str(self._ignore_path) if self._ignore_path else None)

        # Convert mode file pickers (hidden by default)
        self._conv_fp = tk.Frame(fp_container, bg=BG)
        self._convert_in_var  = tk.StringVar(value="尚未選取")
        self._convert_out_var = tk.StringVar(value="尚未選取")
        self._conv_in_row  = self._make_file_row(
            self._conv_fp, "📊 掃描報告:", self._convert_in_var, self._pick_convert_in, row=0)
        self._conv_out_row = self._make_file_row(
            self._conv_fp, "💾 Ignore 輸出:", self._convert_out_var, self._pick_convert_out,
            row=1, is_save=True)
        self._bind_row_tooltip(self._conv_in_row,  lambda: str(self._convert_in_path)  if self._convert_in_path  else None)
        self._bind_row_tooltip(self._conv_out_row, lambda: str(self._convert_out_path) if self._convert_out_path else None)

        # Diff mode file pickers (hidden by default)
        self._diff_fp = tk.Frame(fp_container, bg=BG)
        self._zip_old_var  = tk.StringVar(value="尚未選取")
        self._diff_out_var = tk.StringVar(value="尚未選取")
        self._diff_old_row = self._make_file_row(
            self._diff_fp, "📦 舊版 Zip:", self._zip_old_var, self._pick_zip_old, row=0)
        self._diff_new_row = self._make_file_row(
            self._diff_fp, "📦 新版 Zip:", self._zip_var, self._pick_zip, row=1)
        self._diff_out_row = self._make_file_row(
            self._diff_fp, "💾 輸出 Excel:", self._diff_out_var, self._pick_diff_out,
            row=2, is_save=True)
        self._bind_row_tooltip(self._diff_old_row, lambda: str(self._zip_old_path) if self._zip_old_path else None)
        self._bind_row_tooltip(self._diff_new_row, lambda: str(self._zip_path)     if self._zip_path     else None)
        self._bind_row_tooltip(self._diff_out_row, lambda: str(self._diff_out_path) if self._diff_out_path else None)

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=(10, 8))

        # Run buttons (left column bottom)
        run_row = tk.Frame(left, bg=BG)
        run_row.pack(fill="x")
        _RUN_BG     = "#89b4fa"
        _RUN_BG_ACT = "#74a8e6"
        self._run_btn = tk.Button(run_row, text="▶  Run",
                                  font=("Microsoft JhengHei UI", 14, "bold"),
                                  bg=_RUN_BG, fg="white",
                                  activebackground=_RUN_BG_ACT,
                                  activeforeground="white", relief="flat",
                                  padx=20, pady=8, cursor="hand2",
                                  command=self._run)
        self._run_btn.pack(side="left", padx=(0, 8))
        self._scan_run_btn = tk.Button(run_row, text="▶  開始掃描",
                                       font=("Microsoft JhengHei UI", 14, "bold"),
                                       bg=_RUN_BG, fg="white",
                                       activebackground=_RUN_BG_ACT,
                                       activeforeground="white", relief="flat",
                                       padx=20, pady=8, cursor="hand2",
                                       command=self._run_scan)
        self._convert_run_btn = tk.Button(run_row, text="▶  轉換",
                                          font=("Microsoft JhengHei UI", 14, "bold"),
                                          bg=_RUN_BG, fg="white",
                                          activebackground=_RUN_BG_ACT,
                                          activeforeground="white", relief="flat",
                                          padx=20, pady=8, cursor="hand2",
                                          command=self._run_convert)
        self._diff_run_btn = tk.Button(run_row, text="▶  比對",
                                       font=("Microsoft JhengHei UI", 14, "bold"),
                                       bg=_RUN_BG, fg="white",
                                       activebackground=_RUN_BG_ACT, activeforeground="white",
                                       relief="flat", padx=20, pady=8,
                                       cursor="hand2", command=self._run_diff)
        self._cancel_btn = tk.Button(run_row, text="■  取消",
                                     font=("Microsoft JhengHei UI", 14, "bold"),
                                     bg="#f38ba8", fg="white",
                                     activebackground="#d6607c",
                                     activeforeground="white", relief="flat",
                                     padx=20, pady=8, cursor="hand2",
                                     command=self._cancel_run)
        # Don't pack scan_run_btn, convert_run_btn, or cancel_btn here — _set_mode / _finish_run controls visibility

        # ── RIGHT COLUMN ──────────────────────────────────────────────────────
        right = tk.Frame(body, bg="#27273a")
        right.grid(row=0, column=1, sticky="nsew", padx=(0, 12), pady=12)

        # ── Excel panel (right) ───────────────────────────────────────────────
        self._excel_panel = tk.Frame(right, bg="#27273a")
        self._excel_panel.pack(fill="both", expand=True, padx=12, pady=10)

        tk.Label(self._excel_panel, text="🔍  快速查詢",
                 font=("Microsoft JhengHei UI", 12, "bold"), fg="#a6adc8", bg="#27273a").pack(anchor="w")
        tk.Label(self._excel_panel, text="每行一個英文字串，Ctrl+Enter 執行　｜　支援 * 萬用字元",
                 font=("Microsoft JhengHei UI", 10), fg="#6c7086", bg="#27273a").pack(anchor="w", pady=(0, 4))
        self._ql_text = tk.Text(self._excel_panel, height=5, bg="#313244", fg="#ffffff",
                                font=("Microsoft JhengHei UI", 12), relief="flat",
                                insertbackground="white", padx=8, pady=6)
        self._ql_text.pack(fill="x", pady=(0, 6))
        self._ql_text.bind("<Control-Return>", lambda e: self._quick_lookup())
        self._ql_btn = tk.Button(self._excel_panel, text="🔍  快速查詢",
                                 font=("Microsoft JhengHei UI", 12, "bold"),
                                 bg="#313244", fg="#a6adc8",
                                 activebackground="#45475a",
                                 activeforeground="white", relief="flat",
                                 padx=14, pady=7, cursor="hand2",
                                 command=self._quick_lookup)
        self._ql_btn.pack(anchor="w")

        # Quick-lookup results are written directly to the main log box below

        # ── Scan panel (right) ────────────────────────────────────────────────
        self._scan_panel = tk.Frame(right, bg="#27273a")
        # shown/hidden by _set_mode

        # ── Convert panel (right) ─────────────────────────────────────────────
        self._convert_panel = tk.Frame(right, bg="#27273a")
        tk.Label(self._convert_panel, text="📋  轉換掃描報告為 Ignore Excel",
                 font=("Microsoft JhengHei UI", 12, "bold"), fg="#a6adc8", bg="#27273a").pack(anchor="w")
        tk.Label(self._convert_panel,
                 text=(
                     "將語言掃描報告（含 「<語言> Extracted」Sheet）\n"
                     "一鍵轉換成可直接套用的 Ignore Excel。\n\n"
                     "運作規則：\n"
                     "  • 自動偵測所有 「X Extracted」格式的 Sheet\n"
                     "  • 只收錄 Comment = Pass 的列\n"
                     "  • 每個 Sheet 只記錄該 Sheet 自己的語言\n"
                     "    （忽略「語言」欄可能包含的多語言值）\n"
                     "  • 多個 Key 同格（換行分隔）自動拆分\n"
                     "  • 同一 Key 多語言皆 Pass → 合併為同一列"
                 ),
                 font=("Microsoft JhengHei UI", 10), fg="#6c7086", bg="#27273a",
                 justify="left").pack(anchor="w", pady=(6, 0))

        # ── Diff panel (right) ────────────────────────────────────────────────
        self._diff_panel = tk.Frame(right, bg="#27273a")
        tk.Label(self._diff_panel, text="🆚  比對新字串",
                 font=("Microsoft JhengHei UI", 12, "bold"), fg="#a6adc8", bg="#27273a").pack(anchor="w")
        tk.Label(self._diff_panel,
                 text=(
                     "選取舊版與新版翻譯 Zip，一鍵找出新版新增的英文字串。\n\n"
                     "輸出欄位：\n"
                     "  • Module            — 字串所屬模組\n"
                     "  • New Strings (EN)  — 新增的英文原文\n"
                     "  • Key               — 對應的翻譯 Key\n\n"
                     "輸出結果按 Module 排序，方便逐模組確認。\n"
                     "支援 Auto Filter，可快速篩選特定模組。"
                 ),
                 font=("Microsoft JhengHei UI", 10), fg="#6c7086", bg="#27273a",
                 justify="left").pack(anchor="w", pady=(6, 0))

        tk.Label(self._scan_panel, text="選擇要掃描的語言：",
                 font=("Microsoft JhengHei UI", 10, "bold"), fg="#a6adc8", bg="#27273a"
                 ).pack(anchor="w", pady=(0, 6))

        self._scan_lang_vars: dict[str, tk.BooleanVar] = {}
        cb_frame = tk.Frame(self._scan_panel, bg="#27273a")
        cb_frame.pack(fill="x", pady=(0, 4))

        app  = self._app_var.get()
        cfg  = APP_CONFIGS.get(app, {}) or {}
        langs = [(code, lang) for code, lang in cfg.items() if lang != "en"]
        cols = 4
        for idx, (code, lang) in enumerate(langs):
            var = tk.BooleanVar(value=False)
            self._scan_lang_vars[lang] = var
            tk.Checkbutton(cb_frame, text=code, variable=var,
                           font=("Microsoft JhengHei UI", 9), fg="#cdd6f4", bg="#27273a",
                           activebackground="#27273a", activeforeground="#cba6f7",
                           selectcolor="#45475a", relief="flat", cursor="hand2"
                           ).grid(row=idx // cols, column=idx % cols,
                                  padx=6, pady=2, sticky="w")

        sel_row = tk.Frame(self._scan_panel, bg="#27273a")
        sel_row.pack(anchor="w", pady=(4, 8))
        tk.Button(sel_row, text="全選", font=("Microsoft JhengHei UI", 10),
                  bg="#45475a", fg="white", relief="flat", padx=8, pady=2,
                  cursor="hand2",
                  command=lambda: [v.set(True) for v in self._scan_lang_vars.values()]
                  ).pack(side="left", padx=(0, 6))
        tk.Button(sel_row, text="全不選", font=("Microsoft JhengHei UI", 10),
                  bg="#45475a", fg="white", relief="flat", padx=8, pady=2,
                  cursor="hand2",
                  command=lambda: [v.set(False) for v in self._scan_lang_vars.values()]
                  ).pack(side="left")

        scan_out_f = tk.Frame(self._scan_panel, bg="#27273a")
        scan_out_f.pack(fill="x", pady=(0, 6), padx=(0, 12))
        self._scan_out_var = tk.StringVar(value="尚未選取")
        self._scan_out_row = self._make_file_row(scan_out_f, "💾 輸出報告:",
                            self._scan_out_var, self._pick_scan_out,
                            row=0, is_save=True)
        self._bind_row_tooltip(self._scan_out_row, lambda: str(self._scan_out_path) if self._scan_out_path else None)

        # ── Bottom: progress + log (full width) ───────────────────────────────
        bot = tk.Frame(self._inner, bg=BG)
        bot.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        _pb_style = ttk.Style()
        _pb_style.theme_use("clam")
        _pb_style.configure("Mocha.Horizontal.TProgressbar",
                            troughcolor="#27273a", background="#a6e3a1",
                            bordercolor="#45475a", lightcolor="#a6e3a1", darkcolor="#a6e3a1")
        self._progress = ttk.Progressbar(bot, mode="indeterminate",
                                         style="Mocha.Horizontal.TProgressbar")
        self._progress.pack(fill="x", pady=(0, 2))
        self._progress_label = tk.Label(bot, text="", font=("Microsoft JhengHei UI", 10),
                                        fg="#cba6f7", bg=BG)
        self._progress_label.pack(anchor="w", pady=(0, 4))

        log_f = tk.Frame(bot, bg=BG)
        log_f.pack(fill="both", expand=True)
        self._log_box = tk.Text(log_f, height=1, bg="#11111b", fg="#cdd6f4",
                                font=("Courier", 11), relief="flat",
                                state="disabled", wrap="word",
                                insertbackground="white",
                                selectbackground="#313244", selectforeground="#cdd6f4",
                                inactiveselectbackground="#313244")
        sb = ttk.Scrollbar(log_f, command=self._log_box.yview)
        self._log_box.configure(yscrollcommand=sb.set)
        self._log_box.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._log_box.tag_config("ok",   foreground="#4caf50")
        self._log_box.tag_config("err",  foreground="#f44336")
        self._log_box.tag_config("info", foreground="#a6adc8")
        self._log_box.tag_config("ql",   foreground="#cba6f7")

    @staticmethod
    def _fmt_name(name: str, n: int = 32) -> str:
        return ("…" + name[-(n - 1):]) if len(name) > n else name

    def _bind_tooltip(self, widget, path_fn, anchor_widget=None):
        """Show full path as a floating tooltip on hover; hide on leave.
        anchor_widget: widget whose rootx is used for tooltip x-position.
                       Defaults to widget itself.
        """
        def _show(e):
            path = path_fn()
            if not path:
                return
            tip = tk.Toplevel(self)
            tip.wm_overrideredirect(True)
            tip.configure(bg="#27273a")
            tk.Label(tip, text=path, font=("Microsoft JhengHei UI", 9), fg="#cdd6f4",
                     bg="#27273a", padx=10, pady=5).pack()
            tip.update_idletasks()
            anchor = anchor_widget or widget
            x = anchor.winfo_rootx()
            y = anchor.winfo_rooty() + anchor.winfo_height() + 2
            sw = self.winfo_screenwidth()
            if x + tip.winfo_reqwidth() > sw:
                x = max(0, sw - tip.winfo_reqwidth() - 4)
            tip.geometry(f"+{x}+{y}")
            widget._tip = tip

        def _hide(e):
            tip = getattr(widget, "_tip", None)
            if tip:
                try:
                    tip.destroy()
                except Exception:
                    pass
                widget._tip = None

        widget.bind("<Enter>", _show)
        widget.bind("<Leave>", _hide)

    def _bind_row_tooltip(self, row_widgets, path_fn):
        """Bind tooltip to both the label and filename widget of a file row.
        Tooltip always anchors to lbl (left edge of row) so it appears flush left.
        """
        lbl, val, _ = row_widgets
        self._bind_tooltip(lbl, path_fn)
        self._bind_tooltip(val, path_fn, anchor_widget=lbl)

    def _make_file_row(self, parent, label, var, cmd, row, is_save=False):
        parent.columnconfigure(1, weight=1)   # filename column stretches; button always visible
        lbl = tk.Label(parent, text=label, font=("Microsoft JhengHei UI", 11), fg="#cdd6f4",
                       bg=parent["bg"], width=14, anchor="w")
        lbl.grid(row=row, column=0, sticky="w", pady=4)
        val = tk.Label(parent, textvariable=var, font=("Microsoft JhengHei UI", 9), fg="#a6adc8",
                       bg=parent["bg"], anchor="w")
        val.grid(row=row, column=1, sticky="ew", padx=(4, 0))
        btn = tk.Button(parent, text="選取", font=("Microsoft JhengHei UI", 10),
                        bg="#45475a", fg="white", activebackground="#3d59a1",
                        activeforeground="white", relief="flat", padx=8, pady=2,
                        cursor="hand2", command=cmd)
        btn.grid(row=row, column=2, padx=(4, 0))
        return lbl, val, btn

    def _set_file_row_state(self, widgets, enabled: bool):
        lbl, val, btn = widgets
        if enabled:
            lbl.configure(fg="#cdd6f4")
            val.configure(fg="#a6adc8")
            btn.configure(state="normal", bg="#45475a", fg="white", cursor="hand2")
        else:
            lbl.configure(fg="#6c7086")
            val.configure(fg="#45475a")
            btn.configure(state="disabled", bg="#1e1e2e", fg="#6c7086", cursor="")

    # ── App selector ─────────────────────────────────────────────────────────

    def _select_app(self, app_name: str, init: bool = False):
        if not init:
            # Save current app's paths before switching
            self._save_app_paths()

        self._app_var.set(app_name)
        for name, btn in self._app_buttons.items():
            if name == app_name:
                btn.configure(bg="#cba6f7", fg="#1e1e2e",
                              activebackground="#b4befe", activeforeground="#1e1e2e")
            else:
                btn.configure(bg="#313244", fg="#a6adc8",
                              activebackground="#45475a", activeforeground="#ffffff")
        cfg = APP_CONFIGS.get(app_name)
        if app_name == "Web":
            langs_text = ("  ".join(self._web_target_langs.keys())
                          if self._web_target_langs
                          else "（選取 Zip 後自動偵測語言）")
            self._lang_preview.configure(text=langs_text)
        else:
            self._lang_preview.configure(
                text="  ".join(cfg.keys()) if cfg else "（全部語言）"
            )

        # Disable/enable convert mode for Web
        if hasattr(self, '_mode_buttons'):
            is_web = (app_name == "Web")
            _conv_btn = self._mode_buttons.get("convert")
            if _conv_btn:
                if is_web:
                    if self._mode_var.get() == "convert":
                        self._set_mode("excel")   # switch mode first (re-styles all btns)
                    # Use color + no-op command (state="disabled" ignores bg on flat btns)
                    _conv_btn.configure(
                        bg="#1e1e2e", fg="#3d3d55", cursor="",
                        activebackground="#1e1e2e", activeforeground="#3d3d55",
                        command=lambda: None)
                else:
                    _conv_btn.configure(
                        cursor="hand2",
                        command=lambda: self._set_mode("convert"))
                    # Re-apply proper coloring via _set_mode
                    self._set_mode(self._mode_var.get())

        if not init:
            # Load this app's saved paths (don't clear — restore from config)
            self._zip_path = self._excel_path = self._out_path = self._ignore_path = None
            self._scan_out_path = None
            self._convert_in_path = self._convert_out_path = None
            self._zip_old_path = self._diff_out_path = None
            self._zip_var.set("尚未選取")
            self._excel_var.set("尚未選取")
            self._out_var.set("尚未選取（將放在 Excel 同目錄）")
            self._ignore_var.set("未設定（可選）")
            self._scan_out_var.set("尚未選取")
            self._convert_in_var.set("尚未選取")
            self._convert_out_var.set("尚未選取")
            self._zip_old_var.set("尚未選取")
            self._diff_out_var.set("尚未選取")

            app_cfg = self._cfg.get(app_name, {})
            def _try_set(path_str, var, attr):
                if path_str and Path(path_str).exists():
                    setattr(self, attr, Path(path_str))
                    var.set(self._fmt_name(Path(path_str).name))
            _try_set(app_cfg.get("zip"),    self._zip_var,    "_zip_path")
            _try_set(app_cfg.get("excel"),  self._excel_var,  "_excel_path")
            _try_set(app_cfg.get("out"),    self._out_var,    "_out_path")
            _try_set(app_cfg.get("ignore"), self._ignore_var, "_ignore_path")

            scan_out = app_cfg.get("scan_out")
            if scan_out and Path(scan_out).parent.exists():
                self._scan_out_path = Path(scan_out)
                self._scan_out_var.set(self._fmt_name(Path(scan_out).name))

            _try_set(app_cfg.get("convert_in"),  self._convert_in_var,  "_convert_in_path")
            _try_set(app_cfg.get("convert_out"), self._convert_out_var, "_convert_out_path")
            _try_set(app_cfg.get("zip_old"),     self._zip_old_var,     "_zip_old_path")
            _try_set(app_cfg.get("diff_out"),    self._diff_out_var,    "_diff_out_path")

            self._cfg["last_app"] = app_name
            save_config(self._cfg)
            self._rebuild_scan_checkboxes()

    def _update_web_lang_preview(self):
        """Update lang preview label for Web app after zip is loaded."""
        if self._app_var.get() != "Web": return
        langs = self._web_target_langs or {}
        if langs:
            self._lang_preview.configure(text="  ".join(langs.keys()))
        else:
            self._lang_preview.configure(text="（選取 Zip 後自動偵測語言）")

    def _rebuild_scan_checkboxes(self):
        """Rebuild lang checkboxes when App selection changes."""
        if not hasattr(self, '_scan_lang_vars'): return
        for widget in self._scan_panel.winfo_children():
            if isinstance(widget, tk.Frame):
                for child in widget.winfo_children():
                    child.destroy()
                break
        self._scan_lang_vars.clear()
        app = self._app_var.get()
        if app == "Web":
            cfg   = self._web_target_langs or {}
            langs = [(code, lang) for code, lang in cfg.items()]
        else:
            cfg   = APP_CONFIGS.get(app, {}) or {}
            langs = [(code, lang) for code, lang in cfg.items() if lang != "en"]
        cols  = 4
        frames = [w for w in self._scan_panel.winfo_children() if isinstance(w, tk.Frame)]
        cb_frame = frames[0] if frames else self._scan_panel
        saved_langs = set(self._cfg.get(self._app_var.get(), {}).get("scan_langs", []))
        for idx, (code, lang) in enumerate(langs):
            var = tk.BooleanVar(value=lang in saved_langs)
            self._scan_lang_vars[lang] = var
            tk.Checkbutton(cb_frame, text=code, variable=var,
                           font=("Microsoft JhengHei UI", 9), fg="#cdd6f4", bg="#27273a",
                           activebackground="#27273a", activeforeground="#cba6f7",
                           selectcolor="#45475a", relief="flat", cursor="hand2"
                           ).grid(row=idx // cols, column=idx % cols,
                                  padx=6, pady=2, sticky="w")

    # ── Mode toggle ───────────────────────────────────────────────────────────

    def _set_mode(self, mode: str, init: bool = False):
        self._mode_var.set(mode)
        is_web = hasattr(self, '_app_var') and self._app_var.get() == "Web"
        for val, btn in self._mode_buttons.items():
            if val == "convert" and is_web:
                # Keep convert button visually disabled on Web regardless of mode
                btn.configure(bg="#1e1e2e", fg="#3d3d55", cursor="",
                              activebackground="#1e1e2e", activeforeground="#3d3d55",
                              command=lambda: None)
            elif val == mode:
                btn.configure(bg="#cba6f7", fg="#1e1e2e",
                              activebackground="#b4befe", activeforeground="#1e1e2e")
            else:
                btn.configure(bg="#313244", fg="#6c7086",
                              activebackground="#45475a", activeforeground="#ffffff")

        # Switch run button (guard for init before buttons are created)
        if hasattr(self, '_run_btn') and hasattr(self, '_scan_run_btn'):
            self._run_btn.pack_forget()
            self._scan_run_btn.pack_forget()
            if hasattr(self, '_convert_run_btn'):
                self._convert_run_btn.pack_forget()
            if hasattr(self, '_diff_run_btn'):
                self._diff_run_btn.pack_forget()
            if mode == "excel":
                self._run_btn.pack(side="left", padx=(0, 8))
            elif mode == "scan":
                self._scan_run_btn.pack(side="left")
            elif mode == "convert":
                self._convert_run_btn.pack(side="left")
            else:
                self._diff_run_btn.pack(side="left")

        # Show/hide file picker panels based on mode
        if hasattr(self, '_std_fp') and hasattr(self, '_conv_fp'):
            self._std_fp.pack_forget()
            self._conv_fp.pack_forget()
            if hasattr(self, '_diff_fp'):
                self._diff_fp.pack_forget()
            if mode == "convert":
                self._conv_fp.pack(fill="x")
            elif mode == "diff":
                self._diff_fp.pack(fill="x")
            else:
                self._std_fp.pack(fill="x")

        # Enable/disable 輸入 & 輸出 Excel rows based on mode
        if hasattr(self, '_excel_row') and hasattr(self, '_out_row'):
            excel_mode = (mode == "excel")
            self._set_file_row_state(self._excel_row, excel_mode)
            self._set_file_row_state(self._out_row,   excel_mode)

        if not init:
            # Hide all right panels first
            for _p in ("_excel_panel", "_scan_panel", "_convert_panel", "_diff_panel"):
                if hasattr(self, _p):
                    getattr(self, _p).pack_forget()
            if mode == "excel":
                self._excel_panel.pack(fill="both", expand=True, padx=12, pady=10)
            elif mode == "scan":
                self._scan_panel.pack(fill="x", expand=False, padx=12, pady=(10, 4))
            elif mode == "convert":
                self._convert_panel.pack(fill="x", expand=False, padx=12, pady=(10, 4))
            else:  # diff
                self._diff_panel.pack(fill="x", expand=False, padx=12, pady=(10, 4))

    def _clear_ignore(self):
        self._ignore_path = None
        self._ignore_var.set("未設定（可選）")
        self._save_app_paths()

    # ── Index loader (shared by _run, _run_scan, _quick_lookup) ──────────────

    def _load_index(self, zip_path: Path) -> tuple:
        import tempfile, hashlib
        zip_stat  = zip_path.stat()
        cache_key = hashlib.md5(
            f"{zip_path}|{zip_stat.st_mtime}|{zip_stat.st_size}|{TOOL_VERSION}".encode()
        ).hexdigest()[:12]
        cache_dir  = Path(tempfile.gettempdir()) / "TranslationTool"
        cache_dir.mkdir(exist_ok=True)
        index_path = cache_dir / f"lookup_{zip_path.stem}_{cache_key}.json"

        web_zip = _is_web_zip(zip_path)
        if web_zip:
            # Always update web langs (fast filename scan, even on cache hit)
            detected = _detect_web_langs_from_zip(zip_path)
            self._web_target_langs = detected
            self.after(0, self._rebuild_scan_checkboxes)
            self.after(0, self._update_web_lang_preview)

        if index_path.exists():
            self._log("📂 發現快取索引，直接載入...")
            raw        = json.loads(index_path.read_text("utf-8"))
            index      = {r["en"]: r["projects"] for r in raw}
            norm_index = {_normalize(k): k for k in index}
            self._log(f"✅ 索引載入: {len(index)} 個字串")
            return index, norm_index

        if web_zip:
            index, norm_index, _ = build_web_index(zip_path, self._log)
        else:
            index, norm_index = build_index(zip_path, "en", self._log)
        try:
            out_list = [{"en": k, "projects": v} for k, v in index.items()]
            index_path.write_text(
                json.dumps(out_list, ensure_ascii=False, separators=(",", ":")),
                encoding="utf-8")
            self._log("💾 索引已快取 (下次更快)")
        except Exception:
            self._log("⚠️  快取寫入失敗，不影響本次執行")
        return index, norm_index

    # ── Scan output picker ────────────────────────────────────────────────────

    def _pick_scan_out(self):
        p = filedialog.asksaveasfilename(title="選取輸出報告路徑",
                                         defaultextension=".xlsx",
                                         filetypes=[("Excel files", "*.xlsx")])
        if p:
            self._scan_out_path = Path(p)
            self._scan_out_var.set(self._fmt_name(Path(p).name))
            self._save_app_paths()

    # ── Convert Ignore pickers & run ─────────────────────────────────────────

    def _pick_convert_in(self):
        p = filedialog.askopenfilename(title="選取語言掃描報告",
                                       filetypes=[("Excel files", "*.xlsx *.xls"),
                                                  ("All files", "*.*")])
        if p:
            self._convert_in_path = Path(p)
            self._convert_in_var.set(self._fmt_name(Path(p).name))
            if not self._convert_out_path:
                default = Path(p).parent / (Path(p).stem + "_ignore.xlsx")
                self._convert_out_path = default
                self._convert_out_var.set(self._fmt_name(default.name))
            self._save_app_paths()

    def _pick_convert_out(self):
        p = filedialog.asksaveasfilename(title="選取 Ignore Excel 輸出路徑",
                                          defaultextension=".xlsx",
                                          filetypes=[("Excel files", "*.xlsx")])
        if p:
            self._convert_out_path = Path(p)
            self._convert_out_var.set(self._fmt_name(Path(p).name))
            self._save_app_paths()

    def _ask_file_exists(self, existing: Path) -> Path | None:
        """
        Show a 4-button dialog when output file already exists.
        Returns: resolved Path to use, or None to abort.
        """
        result = {"action": None}
        dlg = tk.Toplevel(self)
        dlg.title("檔案已存在")
        dlg.configure(bg="#1e1e2e")
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text=f"「{existing.name}」已存在，請選擇：",
                 font=("Microsoft JhengHei UI", 11), fg="#cdd6f4", bg="#1e1e2e",
                 padx=20, pady=16).pack()

        btn_f = tk.Frame(dlg, bg="#1e1e2e")
        btn_f.pack(padx=20, pady=(0, 16))

        def _choose(action):
            result["action"] = action
            dlg.destroy()

        for text, action, bg, fg in [
            ("取代",       "overwrite", "#cba6f7", "white"),
            ("自動加序號", "number",    "#45475a", "white"),
            ("取新名稱",   "rename",    "#45475a", "white"),
            ("取消",       "cancel",    "#2d2030", "#f38ba8"),
        ]:
            tk.Button(btn_f, text=text, font=("Microsoft JhengHei UI", 11, "bold"),
                      bg=bg, fg=fg, relief="flat", padx=14, pady=6,
                      cursor="hand2", command=lambda a=action: _choose(a)
                      ).pack(side="left", padx=4)

        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - dlg.winfo_width())  // 2
        y = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{x}+{y}")
        self.wait_window(dlg)

        action = result["action"]
        if action == "overwrite":
            return existing
        elif action == "number":
            base = re.sub(r'_\d+$', '', existing.stem)
            n = 1
            while True:
                candidate = existing.parent / f"{base}_{n}{existing.suffix}"
                if not candidate.exists():
                    return candidate
                n += 1
        elif action == "rename":
            p = filedialog.asksaveasfilename(
                title="另存為",
                initialdir=str(existing.parent),
                initialfile=existing.stem,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            return Path(p) if p else None
        return None   # cancel

    def _run_convert(self):
        if not self._convert_in_path:
            messagebox.showwarning("提示", "請先選取語言掃描報告"); return

        out_path = self._convert_out_path
        if not out_path:
            out_path = self._convert_in_path.parent / (self._convert_in_path.stem + "_ignore.xlsx")
            self._convert_out_path = out_path
            self._convert_out_var.set(self._fmt_name(out_path.name))

        # ── 防呆：輸出檔已存在 ────────────────────────────────────────────────
        if out_path.exists():
            resolved = self._ask_file_exists(out_path)
            if resolved is None:
                return
            out_path = resolved
            self._convert_out_path = out_path
            self._convert_out_var.set(self._fmt_name(out_path.name))
            self._save_app_paths()

        self._cancel_event.clear()
        self._convert_run_btn.configure(state="disabled")
        self._convert_run_btn.pack_forget()
        self._cancel_btn.pack(side="left")
        self._progress.configure(mode="indeterminate")
        self._progress.start(12)
        self._progress_label.configure(text="")
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

        in_path = self._convert_in_path

        def worker():
            try:
                n = convert_scan_to_ignore(in_path, out_path, self._log)
                self._log(f"\n🎉 完成！共 {n} 筆，已儲存至:\n   {out_path}")
                if messagebox.askyesno("完成", f"Ignore Excel 已產生！\n\n{out_path}\n\n是否立即開啟？"):
                    import os
                    if sys.platform == "win32":
                        os.startfile(str(out_path))
                    else:
                        import subprocess
                        subprocess.Popen(["open" if sys.platform == "darwin" else "xdg-open", str(out_path)])
            except Exception as ex:
                import traceback
                self._log(f"❌ 發生錯誤: {ex}", "err")
                self._log(traceback.format_exc(), "err")
                messagebox.showerror("錯誤", str(ex))
            finally:
                self.after(0, self._finish_run, self._convert_run_btn)

        threading.Thread(target=worker, daemon=True).start()

    # ── Scan run ─────────────────────────────────────────────────────────────

    def _run_scan(self):
        if not self._zip_path:
            messagebox.showwarning("提示", "請先選取翻譯 Zip 檔"); return

        selected = [lang for lang, var in self._scan_lang_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("提示", "請至少選擇一個語言"); return

        out_path = getattr(self, "_scan_out_path", None)
        if not out_path:
            out_path = self._zip_path.parent / "scan_report.xlsx"
            self._scan_out_path = out_path
            self._scan_out_var.set(self._fmt_name(out_path.name))

        if out_path.exists():
            resolved = self._ask_file_exists(out_path)
            if resolved is None:
                return
            out_path = resolved
            self._scan_out_path = out_path
            self._scan_out_var.set(self._fmt_name(out_path.name))
            self._save_app_paths()

        self._cancel_event.clear()
        self._scan_run_btn.configure(state="disabled")
        self._scan_run_btn.pack_forget()
        self._cancel_btn.pack(side="left")
        self._progress.configure(mode="indeterminate")
        self._progress.start(12)
        self._progress_label.configure(text="")
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

        zip_path = self._zip_path

        def worker():
            try:
                index, _ = self._load_index(zip_path)
                if self._cancel_event.is_set():
                    self._log("⚠️  已取消"); return
                _scan_app = self._app_var.get()
                if _scan_app == "Web":
                    target_langs = self._web_target_langs or None
                else:
                    target_langs = APP_CONFIGS.get(_scan_app)
                has_module = (_scan_app != "Web")

                ignore_set = set()
                if self._ignore_path and Path(self._ignore_path).exists():
                    ignore_set = load_ignore_list(Path(self._ignore_path), target_langs)
                    self._log(f"🚫 Ignore list: {len(ignore_set)} 筆")

                generate_scan_report(index, out_path, self._log, selected,
                                     target_langs, ignore_set, self._cancel_event,
                                     has_module=has_module)
                if self._cancel_event.is_set(): return
                self._log(f"\n🎉 完成！報告已儲存至:\n   {out_path}")

                if messagebox.askyesno("完成", f"掃描報告已產生！\n\n{out_path}\n\n是否立即開啟？"):
                    import subprocess, os
                    if sys.platform == "win32":   os.startfile(str(out_path))
                    elif sys.platform == "darwin": subprocess.Popen(["open", str(out_path)])
                    else:                          subprocess.Popen(["xdg-open", str(out_path)])

            except Exception as ex:
                import traceback
                self._log(f"❌ 發生錯誤: {ex}", "err")
                self._log(traceback.format_exc(), "err")
                messagebox.showerror("錯誤", str(ex))
            finally:
                self.after(0, self._finish_run, self._scan_run_btn)

        threading.Thread(target=worker, daemon=True).start()

    # ── File pickers ─────────────────────────────────────────────────────────

    def _pick_zip(self):
        p = filedialog.askopenfilename(title="選取翻譯 Zip 檔",
                                       filetypes=[("Zip files", "*.zip"), ("All files", "*.*")])
        if p:
            self._zip_path = Path(p)
            self._zip_var.set(self._fmt_name(Path(p).name))
            if self._app_var.get() == "Web":
                self._web_target_langs = _detect_web_langs_from_zip(self._zip_path)
                self._rebuild_scan_checkboxes()
                self._update_web_lang_preview()
            self._save_app_paths()

    def _pick_excel(self):
        p = filedialog.askopenfilename(title="選取輸入 Excel",
                                       filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if p:
            self._excel_path = Path(p)
            self._excel_var.set(self._fmt_name(Path(p).name))
            if not self._out_path:
                default = self._excel_path.parent / (self._excel_path.stem + "_translated.xlsx")
                self._out_path = default
                self._out_var.set(self._fmt_name(default.name))
            self._save_app_paths()

    def _pick_out(self):
        p = filedialog.asksaveasfilename(title="選取輸出 Excel 路徑",
                                         defaultextension=".xlsx",
                                         filetypes=[("Excel files", "*.xlsx")])
        if p:
            self._out_path = Path(p)
            self._out_var.set(self._fmt_name(Path(p).name))
            self._save_app_paths()

    def _pick_ignore(self):
        p = filedialog.askopenfilename(title="選取 Ignore Excel",
                                       filetypes=[("Excel files", "*.xlsx *.xls"),
                                                  ("All files", "*.*")])
        if p:
            self._ignore_path = Path(p)
            self._ignore_var.set(self._fmt_name(Path(p).name))
            self._save_app_paths()

    def _pick_zip_old(self):
        p = filedialog.askopenfilename(title="選取舊版翻譯 Zip", filetypes=[("Zip", "*.zip")])
        if p:
            self._zip_old_path = Path(p)
            self._zip_old_var.set(self._fmt_name(Path(p).name))
            self._save_app_paths()

    def _pick_diff_out(self):
        p = filedialog.asksaveasfilename(
            title="輸出 Excel 路徑", defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="new_strings.xlsx")
        if p:
            self._diff_out_path = Path(p)
            self._diff_out_var.set(self._fmt_name(Path(p).name))
            self._save_app_paths()

    def _run_diff(self):
        if not self._zip_old_path:
            messagebox.showwarning("缺少檔案", "請選取舊版 Zip 檔"); return
        if not self._zip_path:
            messagebox.showwarning("缺少檔案", "請選取新版 Zip 檔"); return
        if not self._diff_out_path:
            _stem = re.sub(r'[^\w\-]', '_', self._zip_path.stem)[:25]
            default = self._zip_path.parent / f"new_strings_{_stem}.xlsx"
            self._diff_out_path = default
            self._diff_out_var.set(self._fmt_name(default.name))
            self._save_app_paths()

        if self._diff_out_path.exists():
            resolved = self._ask_file_exists(self._diff_out_path)
            if resolved is None:
                return
            self._diff_out_path = resolved
            self._diff_out_var.set(self._fmt_name(resolved.name))
            self._save_app_paths()

        self._cancel_event.clear()
        self._diff_run_btn.configure(state="disabled")
        self._diff_run_btn.pack_forget()
        self._cancel_btn.pack(side="left")
        self._progress.configure(mode="indeterminate")
        self._progress.start(12)
        self._progress_label.configure(text="")
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

        zip_old_path = self._zip_old_path
        zip_new_path = self._zip_path
        out_path     = self._diff_out_path
        cancel       = self._cancel_event

        def _task():
            try:
                self._log(f"🆚 比對新字串")
                self._log(f"   舊版: {zip_old_path.name}")
                self._log(f"   新版: {zip_new_path.name}")
                old_index, _ = self._load_index(zip_old_path)
                if cancel.is_set():
                    self._log("⚠️  已取消"); return
                new_index, _ = self._load_index(zip_new_path)
                if cancel.is_set():
                    self._log("⚠️  已取消"); return
                compare_zips(old_index, new_index, out_path, self._log, cancel,
                             has_module=(self._app_var.get() != "Web"))
                if cancel.is_set(): return
                self._log(f"\n🎉 完成！已儲存至:\n   {out_path}")
                if messagebox.askyesno("完成", f"比對完成！\n\n{out_path}\n\n是否立即開啟？"):
                    import subprocess
                    if sys.platform == "win32":    os.startfile(str(out_path))
                    elif sys.platform == "darwin": subprocess.Popen(["open", str(out_path)])
                    else:                          subprocess.Popen(["xdg-open", str(out_path)])
            except Exception as e:
                import traceback
                self._log(f"❌ 錯誤: {e}", "err")
                self._log(traceback.format_exc(), "err")
                messagebox.showerror("錯誤", str(e))
            finally:
                self.after(0, self._finish_run, self._diff_run_btn)

        threading.Thread(target=_task, daemon=True).start()

    # ── Log ──────────────────────────────────────────────────────────────────

    def _cancel_run(self):
        self._cancel_event.set()
        self._cancel_btn.configure(state="disabled", text="■  取消中…")

    def _finish_run(self, run_btn):
        self._progress.stop()
        self._progress.configure(mode="indeterminate", value=0)
        self._progress_label.configure(text="")
        self._cancel_btn.pack_forget()
        self._cancel_btn.configure(state="normal", text="■  取消")
        run_btn.configure(state="normal")
        if run_btn is self._run_btn:
            run_btn.pack(side="left", padx=(0, 8))
        else:
            run_btn.pack(side="left")

    def _log(self, msg: str, tag: str = None):
        self.after(0, self._log_main, msg, tag)

    def _log_ql(self, msg: str, tag: str = None):
        """Quick-lookup log — merged into main log box."""
        if tag is None:
            tag = "ok" if msg.startswith("✅") else "err" if msg.startswith("❌") else "ql"
        self.after(0, self._log_main, msg, tag)

    def _log_main(self, msg: str, tag: str = None):
        if tag is None:
            tag = "ok" if msg.startswith("✅") else "err" if msg.startswith("❌") else "info"
        self._log_box.configure(state="normal")
        self._log_box.insert("end", msg + "\n", tag)
        self._log_box.see("end")
        self._log_box.configure(state="disabled")
        if "處理第" in msg or "掃描第" in msg:
            self._progress_label.configure(text=msg.strip())
            m = re.search(r'(\d+)\s*/\s*(\d+)', msg)
            if m:
                cur, tot = int(m.group(1)), int(m.group(2))
                if self._progress.cget("mode") == "indeterminate":
                    self._progress.stop()
                    self._progress.configure(mode="determinate")
                self._progress.configure(maximum=tot, value=cur)

    # ── Quick Lookup (#8) ────────────────────────────────────────────────────

    def _quick_lookup(self):
        queries = [l.strip() for l in self._ql_text.get("1.0", "end").splitlines() if l.strip()]
        if not queries:
            messagebox.showwarning("提示", "請輸入要查詢的英文字串"); return
        if not self._zip_path:
            messagebox.showwarning("提示", "請先選取翻譯 Zip 檔"); return

        self._ql_btn.configure(state="disabled")
        self._progress.start(12)
        def worker():
            try:
                index, norm_index = self._load_index(self._zip_path)
                _ql_app = self._app_var.get()
                target_langs = (self._web_target_langs or None
                                if _ql_app == "Web"
                                else APP_CONFIGS.get(_ql_app))
                lang_label   = {v: k for k, v in target_langs.items()} if target_langs else {}
                allowed      = set(target_langs.values()) if target_langs else None

                self._log_ql("─────────────────────────────────────")
                for q in queries:
                    # ── Wildcard search (* glob) ───────────────────────────
                    if '*' in q:
                        import fnmatch
                        hits = [(en, pd) for en, pd in index.items()
                                if fnmatch.fnmatch(en.lower(), q.lower())]
                        if not hits:
                            self._log_ql(f"❌  查無字串: {q!r}", "err")
                        else:
                            self._log_ql(f"✅  {q!r}  → {len(hits)} 筆符合")
                            for en_val, proj_dict in sorted(hits, key=lambda x: x[0].lower()):
                                self._log_ql(f"   📝 {en_val}")
                                for proj, pt in sorted(proj_dict.items()):
                                    for lang, val in sorted(pt.items()):
                                        if lang.startswith('_'): continue
                                        if allowed and lang not in allowed: continue
                                        code   = lang_label.get(lang, lang)
                                        status = "🔴" if (not val or val.strip() == en_val.strip()) else "✅"
                                        self._log_ql(f"      {proj}  [{code}]  {status}  {val[:60]}", "info")
                        continue

                    # ── Normal / fuzzy lookup ──────────────────────────────
                    rec = index.get(q)
                    matched = q
                    if not rec:
                        for k, v in index.items():
                            if k.lower() == q.lower(): rec = v; matched = k; break
                    if not rec:
                        orig = norm_index.get(_normalize(q))
                        if orig: rec = index.get(orig); matched = orig

                    if not rec:
                        self._log_ql(f"❌  查無字串: {q!r}", "err")
                        continue

                    note = f" → 模糊比對: {matched!r}" if matched != q else ""
                    self._log_ql(f"✅  {q!r}{note}")
                    for proj, pt in sorted(rec.items()):
                        for lang, val in sorted(pt.items()):
                            if allowed and lang not in allowed: continue
                            code   = lang_label.get(lang, lang)
                            status = "🔴" if (not val or val.strip() == q.strip()) else "✅"
                            self._log_ql(f"   {proj}  [{code}]  {status}  {val[:60]}", "info")
                self._log_ql("─────────────────────────────────────")

            except Exception as ex:
                self._log_ql(f"❌ 查詢錯誤: {ex}", "err")
            finally:
                self.after(0, lambda: (self._progress.stop(),
                                       self._ql_btn.configure(state="normal")))

        threading.Thread(target=worker, daemon=True).start()

    # ── Run ──────────────────────────────────────────────────────────────────

    def _run(self):
        if not self._zip_path:
            messagebox.showwarning("提示", "請先選取翻譯 Zip 檔"); return
        if not self._excel_path:
            messagebox.showwarning("提示", "請先選取輸入 Excel"); return
        if not self._out_path:
            self._out_path = self._excel_path.parent / (self._excel_path.stem + "_translated.xlsx")
            self._out_var.set(self._fmt_name(self._out_path.name))

        if self._out_path.exists():
            resolved = self._ask_file_exists(self._out_path)
            if resolved is None:
                return
            self._out_path = resolved
            self._out_var.set(self._fmt_name(resolved.name))
            self._save_app_paths()

        self._cancel_event.clear()
        self._run_btn.configure(state="disabled")
        self._run_btn.pack_forget()
        self._cancel_btn.pack(side="left", padx=(0, 8))
        self._progress.configure(mode="indeterminate")
        self._progress.start(12)
        self._progress_label.configure(text="")
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

        zip_path = self._zip_path
        excel_path = self._excel_path
        out_path = self._out_path

        def worker():
            try:
                index, norm_index = self._load_index(zip_path)
                if self._cancel_event.is_set():
                    self._log("⚠️  已取消"); return
                selected_app = self._app_var.get()
                if selected_app == "Web":
                    target_langs = self._web_target_langs or None
                else:
                    target_langs = APP_CONFIGS.get(selected_app)
                has_module = (selected_app != "Web")
                self._log(f"🎯 App: {selected_app}  ({len(target_langs) if target_langs else '全部'} 語言)")

                ignore_set = set()
                if self._ignore_path and Path(self._ignore_path).exists():
                    ignore_set = load_ignore_list(Path(self._ignore_path), target_langs)
                    self._log(f"🚫 Ignore list: {len(ignore_set)} 筆")

                generate_excel(index, norm_index, excel_path, out_path, self._log,
                               target_langs, ignore_set, self._cancel_event,
                               has_module=has_module)
                if self._cancel_event.is_set(): return
                self._log(f"\n🎉 完成！檔案已儲存至:\n   {out_path}")

                if messagebox.askyesno("完成", f"Excel 已產生！\n\n{out_path}\n\n是否立即開啟？"):
                    import subprocess, os
                    if sys.platform == "win32":
                        os.startfile(str(out_path))
                    elif sys.platform == "darwin":
                        subprocess.Popen(["open", str(out_path)])
                    else:
                        subprocess.Popen(["xdg-open", str(out_path)])

            except Exception as ex:
                import traceback
                self._log(f"❌ 發生錯誤: {ex}", "err")
                self._log(traceback.format_exc(), "err")
                messagebox.showerror("錯誤", str(ex))
            finally:
                self.after(0, self._finish_run, self._run_btn)

        threading.Thread(target=worker, daemon=True).start()

    # ── Utils ─────────────────────────────────────────────────────────────────

    def _center(self):
        self.update_idletasks()
        w, h = 880, 750
        self.geometry(f"{w}x{h}")
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    # ── Changelog dialog ──────────────────────────────────────────────────────

    def _show_changelog(self):
        win = tk.Toplevel(self)
        win.title("更新記錄")
        win.configure(bg="#1e1e2e")
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text="更新記錄", font=("Microsoft JhengHei UI", 13, "bold"),
                 fg="#ffffff", bg="#1e1e2e").pack(pady=(16, 6))

        frame = tk.Frame(win, bg="#1e1e2e")
        frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        txt = tk.Text(frame, width=58, height=28,
                      bg="#11111b", fg="#cdd6f4",
                      font=("Microsoft JhengHei UI", 10), relief="flat",
                      wrap="word", state="normal",
                      padx=12, pady=10)
        sb = ttk.Scrollbar(frame, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        txt.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Insert with highlight for version headers
        for line in CHANGELOG.splitlines(keepends=True):
            if re.match(r'^v\d+\.', line):
                txt.insert("end", line, "ver")
            elif line.startswith("─"):
                txt.insert("end", line, "sep")
            elif line.strip().startswith("•"):
                txt.insert("end", line, "item")
            elif line.strip() in ("新功能", "Bug 修正", "效能 / 架構"):
                txt.insert("end", line, "section")
            else:
                txt.insert("end", line)

        txt.tag_config("ver",     foreground="#cba6f7", font=("Microsoft JhengHei UI", 11, "bold"))
        txt.tag_config("sep",     foreground="#45475a")
        txt.tag_config("section", foreground="#a6adc8", font=("Microsoft JhengHei UI", 10, "bold"))
        txt.tag_config("item",    foreground="#cdd6f4")
        txt.configure(state="disabled")

        tk.Button(win, text="關閉", font=("Microsoft JhengHei UI", 10),
                  bg="#45475a", fg="white", relief="flat",
                  padx=20, pady=4, cursor="hand2",
                  command=win.destroy).pack(pady=(0, 14))

        # Centre over main window
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - win.winfo_reqwidth())  // 2
        y = self.winfo_y() + (self.winfo_height() - win.winfo_reqheight()) // 2
        win.geometry(f"+{x}+{y}")

    # ── Auto-update check ─────────────────────────────────────────────────────

    @staticmethod
    def _parse_ver(s: str):
        """Parse 'v1.3.260508_1030' → (1, 3, '260508_1030'). Returns None if unparseable."""
        m = re.search(r'(\d+)\.(\d+)\.(\w+)', s)
        if m:
            return (int(m.group(1)), int(m.group(2)), m.group(3))
        return None

    def _check_for_update(self):
        """Background thread: silently fetch latest_version.txt and compare."""
        try:
            import urllib.request
            with urllib.request.urlopen(_UPDATE_CHECK_URL, timeout=6) as resp:
                latest_raw = resp.read().decode("utf-8").strip()
            cur = self._parse_ver(APP_VERSION)
            new = self._parse_ver(latest_raw)
            if cur and new and new != cur:
                # Trigger if major.minor is newer, OR same major.minor but newer build stamp
                if (new[0], new[1]) > (cur[0], cur[1]) or (
                    (new[0], new[1]) == (cur[0], cur[1]) and new[2] > cur[2]
                ):
                    self.after(0, self._show_update_dialog, latest_raw.strip())
        except Exception:
            pass  # network error or timeout — silently skip

    def _show_update_dialog(self, latest_ver: str):
        import webbrowser
        cur = self._parse_ver(APP_VERSION)
        cur_str = f"v{cur[0]}.{cur[1]}.{cur[2]}" if cur else APP_VERSION
        answer = messagebox.askyesno(
            "🆕 發現新版本",
            f"目前版本：{cur_str}\n最新版本：v{latest_ver}\n\n是否前往 Google Drive 下載新版？",
            icon="info"
        )
        if answer:
            webbrowser.open(_UPDATE_DOWNLOAD_URL)


if __name__ == "__main__":
    App().mainloop()
